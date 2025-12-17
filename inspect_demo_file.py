#!/usr/bin/env python3
"""
Inspect Demo_Budget_With_Risks.xlsx in detail
"""

import openpyxl

print("=" * 80)
print("DEMO FILE INSPECTION")
print("=" * 80)
print()

wb = openpyxl.load_workbook('Demo_Budget_With_Risks.xlsx')

print(f"📊 Sheets: {wb.sheetnames}")
print()

# Sheet 1: 予算管理
ws = wb.active
print(f"📄 Sheet 1: {ws.title}")
print(f"   Rows: {ws.max_row}")
print(f"   Columns: {ws.max_column}")
print()

# Show header
print("   Header:")
headers = [ws.cell(1, col).value for col in range(1, 8)]
print(f"   {headers}")
print()

# Show all data with risk indicators
print("   Data (showing formulas):")
risk_count = 0
for row in range(2, ws.max_row + 1):
    item = ws.cell(row, 1).value
    budget = ws.cell(row, 2).value
    actual = ws.cell(row, 3).value
    diff = ws.cell(row, 4).value
    progress = ws.cell(row, 5).value
    person = ws.cell(row, 6).value
    note = ws.cell(row, 7).value
    
    # Check for risks
    has_risk = False
    risk_type = []
    
    # Check if diff is not a formula (should be =B-C)
    if diff and not str(diff).startswith('='):
        has_risk = True
        risk_type.append("数値(数式なし)")
        risk_count += 1
    
    # Check if progress is not a formula (should be =C/B)
    if progress and not str(progress).startswith('='):
        has_risk = True
        risk_type.append("数値(数式なし)")
        risk_count += 1
    
    # Check for broken references (Z, AA, ZZ columns don't exist)
    if isinstance(diff, str) and ('Z' in diff or 'AA' in diff):
        has_risk = True
        risk_type.append("壊れた参照")
        risk_count += 1
    if isinstance(progress, str) and ('Z' in progress or 'AA' in progress):
        has_risk = True
        risk_type.append("壊れた参照")
        risk_count += 1
    
    marker = "⚠️ " if has_risk else "   "
    risk_info = f" [{', '.join(risk_type)}]" if risk_type else ""
    
    item_str = str(item) if item else ""
    diff_str = str(diff) if diff else ""
    progress_str = str(progress) if progress else ""
    
    print(f"{marker}Row {row:2d}: {item_str:12s} | diff={diff_str:15s} | progress={progress_str:15s}{risk_info}")

print()
print(f"   Total risks in Sheet 1: {risk_count}")
print()

# Sheet 2: 参照データ
if len(wb.sheetnames) > 1:
    ws2 = wb[wb.sheetnames[1]]
    print(f"📄 Sheet 2: {ws2.title}")
    print(f"   Rows: {ws2.max_row}")
    print()
    
    external_links = 0
    for row in range(1, ws2.max_row + 1):
        col1 = ws2.cell(row, 1).value
        col2 = ws2.cell(row, 2).value
        
        # Check for external links
        has_external = False
        if isinstance(col2, str) and '[' in col2 and ']' in col2:
            has_external = True
            external_links += 1
        
        marker = "⚠️ " if has_external else "   "
        print(f"{marker}Row {row}: {col1} | {col2}")
    
    print()
    print(f"   External links in Sheet 2: {external_links}")
    print()

print("=" * 80)
print(f"TOTAL RISKS: {risk_count + external_links if len(wb.sheetnames) > 1 else risk_count}")
print("=" * 80)
