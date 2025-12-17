"""
Period Inference Engine

This module classifies columns as ACTUAL / FORECAST / UNCERTAIN
using a 3-priority system:

Priority 1: Header Keywords (HIGH confidence)
Priority 2: Column Majority Vote (MEDIUM confidence)
Priority 3: Date Fallback (LOW confidence)

CRITICAL: Uses Column Majority Vote, NOT single-row pattern analysis.
This prevents misclassification when aggregate rows have formulas across all periods.
"""

import re
from datetime import datetime, timedelta
from typing import Dict, List, Optional
from src.models import ModelAnalysis, CellInfo
from src.explanation_models import PeriodAttribute


class PeriodInferenceEngine:
    """
    Infers period attributes (ACTUAL/FORECAST/UNCERTAIN) for columns.
    """
    
    def __init__(self):
        """Initialize the engine"""
        pass
    
    def infer_period_attributes(self, model: ModelAnalysis) -> Dict[int, PeriodAttribute]:
        """
        Classify each column as ACTUAL / FORECAST / UNCERTAIN.
        
        Args:
            model: ModelAnalysis object
            
        Returns:
            Dictionary mapping column index to PeriodAttribute
        """
        period_attrs = {}
        
        # Get all columns with data
        columns = self._get_all_columns(model)
        
        for col_idx in columns:
            col_label = self._get_column_header(col_idx, model)
            
            # Priority 1: Header Keywords
            keyword_result = self._check_header_keywords(col_label)
            if keyword_result:
                period_attrs[col_idx] = PeriodAttribute(
                    column_index=col_idx,
                    column_label=col_label,
                    period_type=keyword_result,
                    confidence="HIGH",
                    inference_method="header_keyword",
                    inference_details={"keyword_matched": True}
                )
                continue
            
            # Priority 2: Column Majority Vote
            majority_result = self._column_majority_vote(col_idx, model)
            if majority_result:
                period_attrs[col_idx] = PeriodAttribute(
                    column_index=col_idx,
                    column_label=col_label,
                    period_type=majority_result["period_type"],
                    confidence="MEDIUM",
                    inference_method="column_majority",
                    inference_details=majority_result
                )
                continue
            
            # Priority 3: Date Fallback
            date_result = self._date_fallback(col_label)
            if date_result:
                period_attrs[col_idx] = PeriodAttribute(
                    column_index=col_idx,
                    column_label=col_label,
                    period_type=date_result,
                    confidence="LOW",
                    inference_method="date_fallback",
                    inference_details={"date_parsed": True}
                )
                continue
            
            # Default: UNCERTAIN
            period_attrs[col_idx] = PeriodAttribute(
                column_index=col_idx,
                column_label=col_label,
                period_type="UNCERTAIN",
                confidence="LOW",
                inference_method="default",
                inference_details={}
            )
        
        return period_attrs
    
    def _get_all_columns(self, model: ModelAnalysis) -> List[int]:
        """
        Get all column indices that contain data.
        
        Args:
            model: ModelAnalysis object
            
        Returns:
            List of column indices (0-based)
        """
        from openpyxl.utils import column_index_from_string
        
        columns = set()
        
        for cell_key, cell_info in model.cells.items():
            # Extract column from address
            match = re.match(r'([A-Z]+)(\d+)', cell_info.address)
            if match:
                col_letter = match.group(1)
                col_idx = column_index_from_string(col_letter) - 1  # 0-based
                columns.add(col_idx)
        
        return sorted(list(columns))
    
    def _get_column_header(self, col_idx: int, model: ModelAnalysis) -> str:
        """
        Get header label for a column.
        
        Looks in rows 1-5 for header text.
        
        Args:
            col_idx: Column index (0-based)
            model: ModelAnalysis object
            
        Returns:
            Header label or column letter
        """
        from openpyxl.utils import get_column_letter
        
        col_letter = get_column_letter(col_idx + 1)  # Convert to 1-based
        
        # Check first 5 rows for header
        for row_num in range(1, 6):
            for sheet in model.sheets:
                cell_key = f"{sheet}!{col_letter}{row_num}"
                cell_info = model.cells.get(cell_key)
                
                if cell_info and cell_info.value:
                    # Found a header
                    header_text = str(cell_info.value).strip()
                    if header_text and not header_text.startswith('='):
                        return header_text
        
        # No header found, use column letter
        return col_letter
    
    def _check_header_keywords(self, header: str) -> Optional[str]:
        """
        Check if header contains period keywords.
        
        Args:
            header: Header text
            
        Returns:
            "ACTUAL" or "FORECAST" if keyword found, None otherwise
        """
        if not header:
            return None
        
        header_lower = header.lower()
        
        # ACTUAL keywords (English + Japanese)
        actual_keywords = [
            'act', 'actual', 'actuals',
            '実績', 'じっせき'
        ]
        
        for keyword in actual_keywords:
            if keyword in header_lower:
                return "ACTUAL"
        
        # FORECAST keywords (English + Japanese)
        forecast_keywords = [
            'est', 'estimate', 'plan', 'forecast', 'budget',
            '予測', '計画', '予算', 'よそく', 'けいかく'
        ]
        
        for keyword in forecast_keywords:
            if keyword in header_lower:
                return "FORECAST"
        
        return None
    
    def _column_majority_vote(self, col_idx: int, model: ModelAnalysis) -> Optional[Dict]:
        """
        Determine period type based on column majority vote.
        
        CRITICAL: This is the core logic for period inference.
        
        Algorithm:
        1. Get all constituent cells in the column (exclude headers/totals)
        2. Count hardcoded values vs formulas
        3. Majority hardcode → ACTUAL
        4. Majority formula → FORECAST
        5. Tie → UNCERTAIN
        
        Args:
            col_idx: Column index (0-based)
            model: ModelAnalysis object
            
        Returns:
            Dictionary with period_type and vote counts, or None if no data
        """
        from openpyxl.utils import get_column_letter
        
        col_letter = get_column_letter(col_idx + 1)
        
        # Get all cells in this column (across all sheets)
        constituent_cells = []
        
        for sheet in model.sheets:
            # Scan rows 6-1000 (skip headers in rows 1-5)
            for row_num in range(6, 1001):
                cell_key = f"{sheet}!{col_letter}{row_num}"
                cell_info = model.cells.get(cell_key)
                
                if cell_info and cell_info.value is not None:
                    # Skip total/subtotal rows (common patterns)
                    if self._is_total_row(cell_info, model):
                        continue
                    
                    constituent_cells.append(cell_info)
        
        if not constituent_cells:
            return None
        
        # Count hardcodes vs formulas
        hardcode_count = sum(1 for c in constituent_cells if not c.formula)
        formula_count = len(constituent_cells) - hardcode_count
        
        # Determine period type by majority
        if hardcode_count > formula_count:
            period_type = "ACTUAL"
        elif formula_count > hardcode_count:
            period_type = "FORECAST"
        else:
            period_type = "UNCERTAIN"
        
        return {
            "period_type": period_type,
            "hardcode_count": hardcode_count,
            "formula_count": formula_count,
            "total_cells": len(constituent_cells)
        }
    
    def _is_total_row(self, cell_info: CellInfo, model: ModelAnalysis) -> bool:
        """
        Check if cell is in a total/subtotal row.
        
        Looks for keywords in row headers (columns A-G).
        
        Args:
            cell_info: Cell to check
            model: ModelAnalysis object
            
        Returns:
            True if total row, False otherwise
        """
        # Extract row number
        match = re.match(r'([A-Z]+)(\d+)', cell_info.address)
        if not match:
            return False
        
        row_num = match.group(2)
        
        # Check columns A-G for total keywords
        total_keywords = [
            'total', 'subtotal', 'sum', 'grand total',
            '合計', '小計', '総計', 'ごうけい', 'しょうけい'
        ]
        
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            label_key = f"{cell_info.sheet}!{col_letter}{row_num}"
            label_cell = model.cells.get(label_key)
            
            if label_cell and label_cell.value:
                label_text = str(label_cell.value).lower()
                
                for keyword in total_keywords:
                    if keyword in label_text:
                        return True
        
        return False
    
    def _date_fallback(self, header: str) -> Optional[str]:
        """
        Infer period type from date in header.
        
        If date is more than 3 months ago, assume ACTUAL.
        
        Args:
            header: Header text
            
        Returns:
            "ACTUAL" if old date, None otherwise
        """
        col_date = self._parse_date_from_header(header)
        
        if not col_date:
            return None
        
        # Check if date is more than 3 months ago
        cutoff_date = datetime.now() - timedelta(days=90)
        
        if col_date < cutoff_date:
            return "ACTUAL"
        
        return None
    
    def _parse_date_from_header(self, header: str) -> Optional[datetime]:
        """
        Extract date from header string.
        
        Supports formats:
        - 2024-01, 2024/01
        - Jan 2024, January 2024
        - FY2024, FY 2024
        - Q1 2024, Q1-2024
        
        Args:
            header: Header text
            
        Returns:
            datetime object or None
        """
        if not header:
            return None
        
        # Pattern 1: YYYY-MM or YYYY/MM
        match = re.search(r'(\d{4})[-/](\d{1,2})', header)
        if match:
            year = int(match.group(1))
            month = int(match.group(2))
            try:
                return datetime(year, month, 1)
            except ValueError:
                pass
        
        # Pattern 2: MM-YYYY or MM/YYYY
        match = re.search(r'(\d{1,2})[-/](\d{4})', header)
        if match:
            month = int(match.group(1))
            year = int(match.group(2))
            try:
                return datetime(year, month, 1)
            except ValueError:
                pass
        
        # Pattern 3: Month Name YYYY (e.g., "Jan 2024", "January 2024")
        month_names = {
            'jan': 1, 'january': 1,
            'feb': 2, 'february': 2,
            'mar': 3, 'march': 3,
            'apr': 4, 'april': 4,
            'may': 5,
            'jun': 6, 'june': 6,
            'jul': 7, 'july': 7,
            'aug': 8, 'august': 8,
            'sep': 9, 'september': 9,
            'oct': 10, 'october': 10,
            'nov': 11, 'november': 11,
            'dec': 12, 'december': 12
        }
        
        for month_name, month_num in month_names.items():
            pattern = rf'\b{month_name}\b.*?(\d{{4}})'
            match = re.search(pattern, header.lower())
            if match:
                year = int(match.group(1))
                try:
                    return datetime(year, month_num, 1)
                except ValueError:
                    pass
        
        # Pattern 4: FY YYYY (assume April start)
        match = re.search(r'FY\s*(\d{4})', header, re.IGNORECASE)
        if match:
            year = int(match.group(1))
            try:
                return datetime(year, 4, 1)  # Assume April start
            except ValueError:
                pass
        
        # Pattern 5: Q1-Q4 YYYY (assume quarter start months)
        match = re.search(r'Q([1-4])\s*[-/]?\s*(\d{4})', header, re.IGNORECASE)
        if match:
            quarter = int(match.group(1))
            year = int(match.group(2))
            month = (quarter - 1) * 3 + 1  # Q1=Jan, Q2=Apr, Q3=Jul, Q4=Oct
            try:
                return datetime(year, month, 1)
            except ValueError:
                pass
        
        return None
