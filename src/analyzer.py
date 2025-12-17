"""
Risk detection and health score analyzer

This module provides the ModelAnalyzer class that detects risks in Excel models
and calculates health scores.
"""

from typing import Dict, List, Optional
import networkx as nx
import re

from src.models import ModelAnalysis, RiskAlert, CellInfo, RiskCategory


class ModelAnalyzer:
    """
    Analyzer for detecting risks in Excel models and calculating health scores.
    
    Risk detection includes:
    - Hidden hardcoded values in formulas
    - Circular references
    - Merged cell risks
    - Cross-sheet complexity
    - Timeline gaps
    """
    
    def __init__(self, smart_context=None):
        """
        Initialize the analyzer.
        
        Args:
            smart_context: Optional SmartContextRecovery instance for AI-powered context
        """
        self.smart_context = smart_context
    
    def analyze(self, model: ModelAnalysis, fiscal_start_month: int = 1, 
                allowed_constants: List[float] = None, debug_callback=None) -> ModelAnalysis:
        """
        Run all risk detection rules and calculate health score.
        
        Args:
            model: ModelAnalysis object to analyze
            fiscal_start_month: Starting month of fiscal year (1-12)
            allowed_constants: List of numeric values to exclude from hardcode detection
            debug_callback: Optional callback function for debug logging
            
        Returns:
            Updated ModelAnalysis with risks and health score
        """
        if allowed_constants is None:
            allowed_constants = []
        
        self.debug_callback = debug_callback
        
        risks: List[RiskAlert] = []
        
        # Run all risk detection methods
        risks.extend(self._detect_hidden_hardcodes(model.cells, allowed_constants))
        risks.extend(self._detect_circular_references(model.dependency_graph))
        risks.extend(self._detect_merged_cell_risks(model.cells, model.merged_ranges))
        risks.extend(self._detect_cross_sheet_spaghetti(model.cells))
        risks.extend(self._detect_timeline_gaps(model, fiscal_start_month))
        
        # DIAGNOSTIC SUITE - Advanced Logic Checks (December 2025)
        risks.extend(self._detect_row_inconsistency(model.cells))
        risks.extend(self._detect_value_conflicts(model.cells))
        risks.extend(self._detect_external_links(model.cells))
        risks.extend(self._detect_formula_errors(model.cells))
        
        # Add impact scores to individual risks BEFORE compression
        risks = self._add_impact_scores(risks, model)
        
        # Compress duplicate risks (Plan A) - will sum up impact scores
        risks = self._compress_risks(risks)
        
        # Add contextual labels to risks
        risks = self._add_context_labels(risks, model.cells)
        
        # Calculate quantitative severity for all risks
        risks = self._calculate_quantitative_severity(risks)
        
        # Calculate health score
        health_score = self._calculate_health_score(risks)
        
        # Update model
        model.risks = risks
        model.health_score = health_score
        
        return model
    
    def _log(self, level: str, message: str, details=None):
        """Log message via callback or print"""
        if self.debug_callback:
            self.debug_callback(level, message, details)
        else:
            print(f"[{level}] {message}")
    
    def _detect_hidden_hardcodes(self, cells: Dict[str, CellInfo], 
                                  allowed_constants: List[float] = None) -> List[RiskAlert]:
        """
        Find ALL numeric literals in formulas with tiered severity.
        
        TIERED VISIBILITY:
        - Common constants (0, 1, 12): LOW severity (user can toggle visibility)
        - Unknown values: HIGH severity (always shown)
        
        Philosophy: Detect everything, let the user choose what to see.
        
        Args:
            cells: Dictionary of CellInfo objects
            allowed_constants: List of numeric values to exclude completely
            
        Returns:
            List of RiskAlert objects for hardcoded values
        """
        from openpyxl.formula.tokenizer import Tokenizer, Token
        
        if allowed_constants is None:
            allowed_constants = []
        
        risks = []
        common_constants = {0, 1, 12}  # Common values - LOW severity
        
        for cell_key, cell_info in cells.items():
            if not cell_info.formula:
                continue
            
            try:
                # Ensure formula starts with '='
                formula_str = cell_info.formula if cell_info.formula.startswith('=') else f"={cell_info.formula}"
                
                # Tokenize the formula
                tokenizer = Tokenizer(formula_str)
                tokens = tokenizer.items
                
                # Look for NUMBER tokens (not RANGE tokens)
                hardcoded_values = []
                for token in tokens:
                    if token.type == Token.OPERAND and token.subtype == Token.NUMBER:
                        # This is a numeric literal, not a cell reference
                        try:
                            num = float(token.value)
                            
                            # Skip user-configured allowed constants only
                            if num in allowed_constants:
                                continue
                            
                            # Found a hardcoded value!
                            hardcoded_values.append(token.value)
                            
                        except ValueError:
                            # Not a valid number, skip
                            continue
                
                # If we found any hardcoded values, create ONE alert per cell
                if hardcoded_values:
                    values_str = ", ".join(hardcoded_values)
                    
                    # Determine severity: LOW for common constants, HIGH for others
                    first_value = float(hardcoded_values[0])
                    is_common = (first_value == int(first_value) and int(first_value) in common_constants)
                    severity = "Low" if is_common else "High"
                    
                    risks.append(RiskAlert(
                        risk_type="Hidden Hardcode",
                        severity=severity,
                        sheet=cell_info.sheet,
                        cell=cell_info.address,
                        description=f"Formula contains hardcoded value(s): {values_str}",
                        details={
                            "formula": cell_info.formula,
                            "hardcoded_value": hardcoded_values[0],  # Use first for grouping
                            "all_hardcoded_values": hardcoded_values
                        }
                    ))
            
            except Exception:
                # If tokenization fails, skip this cell
                continue
        
        return risks
    
    def _detect_circular_references(self, graph: nx.DiGraph) -> List[RiskAlert]:
        """
        Use networkx.simple_cycles() to find cycles.
        
        Args:
            graph: Dependency graph
            
        Returns:
            List of RiskAlert objects for circular references
        """
        risks = []
        
        try:
            # Find all cycles in the graph
            cycles = list(nx.simple_cycles(graph))
            
            # Limit to first 100 cycles to prevent overwhelming output
            cycles_to_report = cycles[:100]
            
            for cycle in cycles_to_report:
                # Get the first cell in the cycle for reporting
                first_cell = cycle[0]
                sheet, address = first_cell.split('!')
                
                # Create a readable cycle description
                cycle_str = " → ".join([c.split('!')[1] for c in cycle[:5]])
                if len(cycle) > 5:
                    cycle_str += f" ... ({len(cycle)} cells total)"
                
                risks.append(RiskAlert(
                    risk_type="Circular Reference",
                    severity="Critical",
                    sheet=sheet,
                    cell=address,
                    description=f"Circular reference detected: {cycle_str}",
                    details={
                        "cycle": cycle,
                        "cycle_length": len(cycle)
                    }
                ))
            
            # If there are more than 100 cycles, add a summary alert
            if len(cycles) > 100:
                risks.append(RiskAlert(
                    risk_type="Circular Reference",
                    severity="Critical",
                    sheet="Multiple",
                    cell="Multiple",
                    description=f"100+ circular references detected (showing first 100)",
                    details={
                        "total_cycles": len(cycles)
                    }
                ))
        
        except Exception as e:
            # If cycle detection fails, log it but don't crash
            print(f"Warning: Circular reference detection failed: {str(e)}")
        
        return risks
    
    def _detect_merged_cell_risks(self, cells: Dict[str, CellInfo], 
                                   merged_ranges: Dict[str, List[str]]) -> List[RiskAlert]:
        """
        Check if formulas reference ranges containing merged cells.
        
        Args:
            cells: Dictionary of CellInfo objects
            merged_ranges: Dictionary of merged ranges by sheet
            
        Returns:
            List of RiskAlert objects for merged cell risks
        """
        risks = []
        
        if not merged_ranges:
            return risks
        
        # Helper function to check if two ranges overlap
        def ranges_overlap(range1: str, range2: str) -> bool:
            """Check if two Excel ranges overlap (simplified check)"""
            try:
                # For MVP, we'll do a simple string comparison
                # A full implementation would parse cell coordinates and check overlap
                # For now, just check if the ranges are exactly the same
                return range1 == range2
            except:
                return False
        
        # Check each cell with a formula
        for cell_key, cell_info in cells.items():
            if not cell_info.formula:
                continue
            
            # Check if this cell references any merged ranges
            # Look for range references in dependencies
            for dep in cell_info.dependencies:
                if ':' in dep:  # This is a range reference
                    dep_sheet, dep_range = dep.split('!')
                    
                    # Check if this sheet has merged ranges
                    if dep_sheet not in merged_ranges:
                        continue
                    
                    # Check if the referenced range overlaps with any merged ranges
                    overlapping_merged = []
                    for merged_range in merged_ranges[dep_sheet]:
                        if ranges_overlap(dep_range, merged_range):
                            overlapping_merged.append(merged_range)
                    
                    # Only report if there's actual overlap
                    if overlapping_merged:
                        risks.append(RiskAlert(
                            risk_type="Merged Cell Risk",
                            severity="Medium",
                            sheet=cell_info.sheet,
                            cell=cell_info.address,
                            description=f"Formula references merged cell range: {dep}",
                            details={
                                "formula": cell_info.formula,
                                "referenced_range": dep,
                                "overlapping_merged_ranges": overlapping_merged
                            }
                        ))
                        # Only report once per cell
                        break
        
        return risks
    
    def _detect_cross_sheet_spaghetti(self, cells: Dict[str, CellInfo]) -> List[RiskAlert]:
        """
        Count distinct external sheet references (>2 = risk).
        
        Args:
            cells: Dictionary of CellInfo objects
            
        Returns:
            List of RiskAlert objects for cross-sheet complexity
        """
        risks = []
        
        for cell_key, cell_info in cells.items():
            if not cell_info.formula or not cell_info.dependencies:
                continue
            
            # Count distinct external sheets referenced
            external_sheets = set()
            current_sheet = cell_info.sheet
            
            for dep in cell_info.dependencies:
                dep_sheet = dep.split('!')[0]
                if dep_sheet != current_sheet:
                    external_sheets.add(dep_sheet)
            
            # If more than 2 external sheets, it's spaghetti
            if len(external_sheets) > 2:
                risks.append(RiskAlert(
                    risk_type="Cross-Sheet Spaghetti",
                    severity="Low",
                    sheet=cell_info.sheet,
                    cell=cell_info.address,
                    description=f"Formula references {len(external_sheets)} external sheets",
                    details={
                        "formula": cell_info.formula,
                        "external_sheets": list(external_sheets),
                        "sheet_count": len(external_sheets)
                    }
                ))
        
        return risks
    
    def _detect_timeline_gaps(self, model: ModelAnalysis, 
                              fiscal_start_month: int) -> List[RiskAlert]:
        """
        Identify missing intervals in date-based row headers.
        
        Args:
            model: ModelAnalysis object
            fiscal_start_month: Starting month of fiscal year
            
        Returns:
            List of RiskAlert objects for timeline gaps
        """
        # Simplified implementation for MVP
        # Full implementation would parse dates and check for gaps
        # For now, we'll skip this complex feature
        return []
    
    def _compress_risks(self, risks: List[RiskAlert]) -> List[RiskAlert]:
        """
        Compress duplicate risks into grouped alerts with spatial proximity checking.
        
        APPLIES TO ALL RISK TYPES (Hardcodes, External Links, Inconsistent Formulas, etc.)
        
        Groups risks by: Risk Type + Sheet + Spatial Proximity
        
        CRITICAL: Don't group F4 (Driver) with F24 (Calculation) across 20-row gap.
        This hides model structure.
        
        Args:
            risks: List of detected risks
            
        Returns:
            Compressed list of risks
        """
        from collections import defaultdict
        
        # Group risks by type and sheet
        risk_groups = defaultdict(list)
        
        for risk in risks:
            # Create a grouping key based on risk type and sheet
            # For hardcodes, also group by value
            if risk.risk_type == "Hidden Hardcode":
                hardcoded_value = risk.details.get("hardcoded_value", "")
                key = (risk.risk_type, risk.sheet, hardcoded_value)
            elif risk.risk_type in ["External Link", "Phantom Link"]:
                # Group external links by sheet (they're all the same type)
                key = (risk.risk_type, risk.sheet, "external_link")
            elif risk.risk_type in ["Inconsistent Formula"]:
                # Group inconsistent formulas by sheet AND row
                # Extract row number from cell address
                row_num = None
                if risk.cell:
                    match = re.match(r'([A-Z]+)(\d+)', risk.cell.split(',')[0].split(':')[0].strip())
                    if match:
                        row_num = match.group(2)
                key = (risk.risk_type, risk.sheet, f"row_{row_num}" if row_num else "inconsistent")
            elif risk.risk_type in ["Inconsistent Value", "Value Conflict"]:
                # Group value conflicts by sheet
                key = (risk.risk_type, risk.sheet, "conflict")
            else:
                # Other risk types: don't compress (keep unique)
                key = (risk.risk_type, risk.sheet, risk.cell, id(risk))
            
            risk_groups[key].append(risk)
        
        # Create compressed risks with spatial proximity checking
        compressed = []
        for key, group in risk_groups.items():
            if len(group) == 1:
                # Single risk, no compression needed
                compressed.append(group[0])
            else:
                # Multiple risks - apply spatial proximity grouping
                # Sort by row number for spatial analysis
                sorted_group = sorted(group, key=lambda r: self._extract_row_number(r.cell))
                
                # Split into spatially proximate clusters
                clusters = self._split_by_spatial_proximity(sorted_group)
                
                # Compress each cluster separately
                for cluster in clusters:
                    compressed.append(self._create_compressed_risk(cluster))
        
        return compressed
    
    def _extract_row_number(self, cell_address: str) -> int:
        """Extract row number from cell address (e.g., 'F24' -> 24)"""
        match = re.match(r'[A-Z]+(\d+)', cell_address)
        return int(match.group(1)) if match else 0
    
    def _extract_row_col(self, cell_address: str) -> tuple:
        """
        Extract row and column numbers from cell address.
        
        FIX 2: "Long Distance" Bug - Need both row AND column for proximity check
        
        Args:
            cell_address: Cell address (e.g., 'F24', 'BN13')
            
        Returns:
            Tuple of (row_number, col_number) e.g., ('F24' -> (24, 6))
        """
        from openpyxl.utils import column_index_from_string
        
        match = re.match(r'([A-Z]+)(\d+)', cell_address)
        if match:
            col_letter = match.group(1)
            row_num = int(match.group(2))
            col_num = column_index_from_string(col_letter)
            return row_num, col_num
        return 0, 0
    
    def _split_by_spatial_proximity(self, sorted_risks: List[RiskAlert], max_gap: int = 1) -> List[List[RiskAlert]]:
        """
        Split risks into spatially proximate clusters.
        
        FIX 2: "Long Distance" Bug - Check BOTH row AND column proximity
        Rule: If gap > 1 row OR column, split the group.
        This prevents grouping F4 with BN4 (same row, far columns).
        
        Args:
            sorted_risks: Risks sorted by row number
            max_gap: Maximum gap to allow in same cluster (default: 1)
            
        Returns:
            List of risk clusters
        """
        if not sorted_risks:
            return []
        
        clusters = []
        current_cluster = [sorted_risks[0]]
        prev_row, prev_col = self._extract_row_col(sorted_risks[0].cell)
        
        for risk in sorted_risks[1:]:
            curr_row, curr_col = self._extract_row_col(risk.cell)
            
            # FIX 2: Check BOTH row and column gaps
            # Only group if cells are touching (gap <= 1 in BOTH dimensions)
            row_gap = abs(curr_row - prev_row)
            col_gap = abs(curr_col - prev_col)
            
            # STRICT RULE: Both gaps must be <= 1 (neighbors only)
            # F4, F5 = OK (row gap 1, col gap 0)
            # F4, G4 = OK (row gap 0, col gap 1)
            # F4, BN4 = NOT OK (row gap 0, col gap 60+)
            # F4, F8 = NOT OK (row gap 4, col gap 0)
            if row_gap <= max_gap and col_gap <= max_gap:
                # Close enough - add to current cluster
                current_cluster.append(risk)
            else:
                # Too far - start new cluster
                clusters.append(current_cluster)
                current_cluster = [risk]
            
            prev_row, prev_col = curr_row, curr_col
        
        # Add final cluster
        clusters.append(current_cluster)
        
        return clusters
    
    def _create_compressed_risk(self, group: List[RiskAlert]) -> RiskAlert:
        """Create a compressed risk from a group of risks (works for ALL risk types)"""
        first_risk = group[0]
        cells = [r.cell for r in group]
        
        # Create range description
        # Always use range format for multiple cells (cleaner display)
        if len(cells) == 1:
            location_str = cells[0]
        elif len(cells) == 2:
            location_str = f"{cells[0]}, {cells[1]}"
        else:
            # Use Excel-style range notation for 3+ cells
            location_str = f"{cells[0]}:{cells[-1]}"
        
        # Sum up impact counts from all risks in the group
        total_impact = sum(risk.details.get("impact_count", 0) for risk in group)
        
        # Create description based on risk type
        if first_risk.risk_type == "Hidden Hardcode":
            hardcoded_value = first_risk.details.get("hardcoded_value", "")
            new_description = f"Hardcoded value '{hardcoded_value}' ({len(group)} instances)"
            details = {
                "hardcoded_value": hardcoded_value,
                "instance_count": len(group),
                "cells": cells,
                "impact_count": total_impact
            }
        elif first_risk.risk_type in ["External Link", "Phantom Link"]:
            new_description = f"External link detected ({len(group)} instances)"
            details = {
                "instance_count": len(group),
                "cells": cells,
                "impact_count": total_impact
            }
        elif first_risk.risk_type in ["Inconsistent Formula"]:
            new_description = f"Inconsistent formula pattern ({len(group)} instances)"
            details = {
                "instance_count": len(group),
                "cells": cells,
                "impact_count": total_impact,
                "formula": first_risk.details.get("formula", ""),  # Preserve formula from first risk
                "pattern": first_risk.details.get("pattern", ""),
                "likelihood_assessment": first_risk.details.get("likelihood_assessment", "")
            }
        elif first_risk.risk_type in ["Inconsistent Value", "Value Conflict"]:
            new_description = f"Conflicting values detected ({len(group)} instances)"
            details = {
                "instance_count": len(group),
                "cells": cells,
                "impact_count": total_impact
            }
        else:
            # Generic compression for other types
            new_description = f"{first_risk.risk_type} ({len(group)} instances)"
            details = {
                "instance_count": len(group),
                "cells": cells,
                "impact_count": total_impact
            }
        
        # Create compressed risk
        return RiskAlert(
            risk_type=first_risk.risk_type,
            severity=first_risk.severity,
            sheet=first_risk.sheet,
            cell=location_str,
            description=new_description,
            details=details
        )
    
    def _get_leftmost_cell(self, cells: List[str]) -> str:
        """
        Get the leftmost cell from a list of cell addresses.
        
        This is critical for range risks - we want to use the leftmost cell
        for context lookup so we can find item names in columns A-G.
        
        Args:
            cells: List of cell addresses (e.g., ["H22", "K22"])
            
        Returns:
            Leftmost cell address (e.g., "H22")
        """
        from openpyxl.utils import column_index_from_string
        
        if not cells:
            return ""
        
        if len(cells) == 1:
            return cells[0]
        
        # Parse each cell and find the one with lowest column number
        leftmost = cells[0]
        leftmost_col_num = 999999
        
        for cell in cells:
            # Parse cell address (e.g., "H22" -> col="H", row=22)
            match = re.match(r'^([A-Z]+)(\d+)$', cell)
            if match:
                col_letter = match.group(1)
                col_num = column_index_from_string(col_letter)
                
                if col_num < leftmost_col_num:
                    leftmost_col_num = col_num
                    leftmost = cell
        
        return leftmost
    
    def _is_poor_quality_label(self, text: str, verbose: bool = False) -> bool:
        """
        Validate if a context label is meaningful or garbage.
        
        Returns True if label is poor quality (triggers AI recovery).
        Returns False if label is acceptable.
        
        Poor quality patterns:
        1. Starts with = (formula debris)
        2. Contains math operators without spaces (formula fragments)
        3. Cell Address Pattern (e.g., "E92", "AA1", "B123")
        4. Generic Stopwords (English/Japanese)
        5. Symbols/Numeric Only (e.g., "-", "0", "123", "---")
        6. Too short (< 2 chars) or too long (> 50 chars)
        
        Args:
            text: Context label to validate
            verbose: Enable verbose logging for debugging
            
        Returns:
            True if poor quality, False if acceptable
        """
        if not text or not text.strip():
            if verbose:
                print(f"[FILTER] Label is empty -> Poor Quality")
            return True
        
        text = text.strip()
        
        # Pattern 1: Starts with = (formula debris like "=(D18*E18)")
        if text.startswith('='):
            if verbose:
                print(f"[FILTER] Label '{text}' starts with = -> Poor Quality")
            return True
        
        # Pattern 2: Contains math operators without spaces (formula fragments)
        # Check for +, *, /, - with no spaces around them
        if re.search(r'[+*/]', text) and ' ' not in text:
            if verbose:
                print(f"[FILTER] Label '{text}' contains math operators -> Poor Quality")
            return True
        
        # Pattern 3: Cell Address Pattern (e.g., "E92", "AA1", "B123", "F24", "BN13")
        # Case insensitive to catch "f24" as well
        if re.match(r'^[A-Za-z]+[0-9]+$', text):
            if verbose:
                print(f"[FILTER] Label '{text}' is cell address -> Poor Quality")
            return True
        
        # Pattern 4: Generic Stopwords (English/Japanese)
        stopwords = {
            # English
            "Total", "Sum", "Subtotal", "Check", "Val", "Value",
            "Amount", "Number", "Item", "Row", "Column",
            # Japanese
            "合計", "小計", "計", "チェック", "検証", "値", "金額"
        }
        if text in stopwords:
            if verbose:
                print(f"[FILTER] Label '{text}' is generic stopword -> Poor Quality")
            return True
        
        # Pattern 5: Symbols/Numeric Only (e.g., "-", "0", "123", "---")
        if re.match(r'^[-0-9\s]+$', text):
            if verbose:
                print(f"[FILTER] Label '{text}' is symbols/numeric only -> Poor Quality")
            return True
        
        # Pattern 6: Too short or too long
        if len(text) < 2 or len(text) > 50:
            if verbose:
                print(f"[FILTER] Label '{text}' length {len(text)} out of range [2-50] -> Poor Quality")
            return True
        
        if verbose:
            print(f"[FILTER] Label '{text}' passed all checks -> Good Quality")
        
        return False
    
    def _add_context_labels(self, risks: List[RiskAlert], cells: Dict[str, CellInfo]) -> List[RiskAlert]:
        """
        Add row and column context labels to risks with AI recovery for poor quality labels.
        
        Args:
            risks: List of risks to add context to
            cells: Dictionary of all cells
            
        Returns:
            Updated list of risks with context labels
        """
        ai_calls = 0
        ai_successes = 0
        empty_contexts = 0
        poor_quality_contexts = 0
        
        # DEBUG: Check if AI is configured
        print(f"\n[DEBUG] _add_context_labels called with {len(risks)} risks")
        print(f"[DEBUG] smart_context configured: {self.smart_context is not None}")
        if self.smart_context:
            print(f"[DEBUG] smart_context enabled: {self.smart_context.enabled}")
        
        risk_count = 0
        for risk in risks:
            risk_count += 1
            verbose = risk_count <= 5  # Verbose logging for first 5 risks
            
            # Handle compressed risks (ranges like "D5...K5" or "H22:K22")
            # CRITICAL FIX: Extract LEFTMOST cell from range for context lookup
            # This ensures we can find item names in columns A-G even for far-right ranges
            cell_for_context = risk.cell
            if '...' in risk.cell:
                # Extract cells from range "D5...K5" -> ["D5", "K5"]
                cells_in_range = [c.strip() for c in risk.cell.split('...')]
                # Pick leftmost cell (lowest column number)
                cell_for_context = self._get_leftmost_cell(cells_in_range)
            elif ':' in risk.cell:
                # Extract cells from range "H22:K22" -> ["H22", "K22"]
                cells_in_range = [c.strip() for c in risk.cell.split(':')]
                # Pick leftmost cell (lowest column number)
                cell_for_context = self._get_leftmost_cell(cells_in_range)
            elif ',' in risk.cell:
                # Extract cells from list "E189, F189" -> ["E189", "F189"]
                cells_in_range = [c.strip() for c in risk.cell.split(',')]
                # Pick leftmost cell (lowest column number)
                cell_for_context = self._get_leftmost_cell(cells_in_range)
            
            # Get row and column labels (use first cell from range)
            row_label, col_label = self._get_context_labels(risk.sheet, cell_for_context, cells)
            
            # NUCLEAR TRIM: Handle Japanese full-width spaces (\u3000) and all whitespace
            if row_label:
                # Replace Japanese full-width space with regular space, then strip
                row_label = row_label.replace('\u3000', ' ').strip()
                if not row_label:  # If empty after nuclear trim
                    row_label = None
            
            if col_label:
                col_label = col_label.replace('\u3000', ' ').strip()
                if not col_label:
                    col_label = None
            
            if verbose:
                print(f"\n[VERBOSE] Risk #{risk_count}: {risk.sheet}!{risk.cell}")
                if cell_for_context != risk.cell:
                    print(f"[VERBOSE] Using first cell for context: {cell_for_context}")
                print(f"[VERBOSE] Row label after nuclear trim: '{row_label}'")
                print(f"[VERBOSE] Col label after nuclear trim: '{col_label}'")
            
            # ENFORCE ROW LABEL PRIORITY: Row label is mandatory
            # Knowing "When" (2025) without "What" (Revenue) is useless
            is_empty = not row_label or row_label == ""
            is_poor = False
            
            if is_empty:
                empty_contexts += 1
                print(f"[DEBUG] Row label missing for {risk.sheet}!{risk.cell}. Triggering AI.")
                if verbose:
                    print(f"[VERBOSE] Label is empty or whitespace-only")
            else:
                is_poor = self._is_poor_quality_label(row_label, verbose=verbose)
                if is_poor:
                    poor_quality_contexts += 1
                    print(f"[AI] Poor quality context '{row_label}' detected for {risk.sheet}!{risk.cell}")
            
            # SMART AI TRIGGER: Enable for critical risk types
            # Priority: Hidden Hardcode, Inconsistent Formula, Inconsistent Value
            critical_risk_types = ["Hidden Hardcode", "Inconsistent Formula", "Inconsistent Value", "Row Inconsistency"]
            should_use_ai = (is_empty or is_poor) and risk.risk_type in critical_risk_types
            
            if self.smart_context and self.smart_context.enabled and should_use_ai:
                ai_calls += 1
                print(f"[AI] Attempting recovery for {risk.sheet}!{risk.cell}")
                ai_label = self.smart_context.recover_context(risk.sheet, cell_for_context, cells)
                
                # Clean AI response - reject "NONE" as invalid
                if ai_label:
                    ai_label = ai_label.strip()
                    # Reject "NONE" - AI couldn't find a label
                    if ai_label.upper() == "NONE":
                        ai_label = None
                
                if ai_label:
                    print(f"[AI] ✓ Recovered: '{ai_label}'")
                    row_label = ai_label
                    ai_successes += 1
                else:
                    print(f"[AI] ✗ Recovery failed or returned NONE")
                    # FALLBACK MECHANISM: Coordinate placeholder
                    if not row_label or row_label == "":
                        # Extract row number from cell address (use first cell from range)
                        row_match = re.match(r'^[A-Z]+(\d+)$', cell_for_context)
                        if row_match:
                            row_num = row_match.group(1)
                            row_label = f"[Unknown Row {row_num}]"
                            print(f"[FALLBACK] Using coordinate placeholder: '{row_label}'")
            
            risk.row_label = row_label
            risk.col_label = col_label
        
        print(f"\n[DEBUG] Summary: {empty_contexts} empty, {poor_quality_contexts} poor quality, {ai_calls} AI calls")
        if ai_calls > 0:
            print(f"[AI] Summary: {ai_successes}/{ai_calls} successful recoveries")
        
        return risks
    
    def _add_impact_scores(self, risks: List[RiskAlert], model: ModelAnalysis) -> List[RiskAlert]:
        """
        Add impact scores (dominance) to all risks.
        
        Args:
            risks: List of RiskAlert objects
            model: ModelAnalysis object with dependency graph
            
        Returns:
            Updated list of RiskAlert objects with impact_count in details
        """
        for risk in risks:
            # Calculate dominance (number of dependent cells)
            cell_address = f"{risk.sheet}!{risk.cell}"
            dominance = self._calculate_dominance(model, cell_address)
            
            # Add to details
            risk.details["impact_count"] = dominance
        
        return risks
    
    def _calculate_quantitative_severity(self, risks: List[RiskAlert]) -> List[RiskAlert]:
        """
        Calculate quantitative severity based on risk scoring formula.
        
        Formula:
        Risk Score = Category Weight × Impact × Error Probability Modifier
        
        Category Weights:
        - Fatal Error: 1.0 (100% - already broken)
        - Integrity Risk: 0.7 (70% - likely error)
        - Structural Debt: 0.3 (30% - future risk)
        
        Error Probability Modifiers (for Inconsistent Formula):
        - Likely error: 1.0 (100%)
        - Uncertain: 0.8 (80%)
        - Possibly intentional: 0.5 (50%)
        - Likely intentional: 0.3 (30%)
        
        Severity Thresholds:
        - Critical: >= 50
        - High: >= 20
        - Medium: >= 5
        - Low: < 5
        
        Args:
            risks: List of RiskAlert objects
            
        Returns:
            Updated list with quantitative severity
        """
        from src.models import RiskCategory
        
        # Category weights
        category_weights = {
            RiskCategory.FATAL_ERROR: 1.0,
            RiskCategory.INTEGRITY_RISK: 0.7,
            RiskCategory.STRUCTURAL_DEBT: 0.3
        }
        
        for risk in risks:
            # Step 1: Determine category
            category = classify_risk(risk, risks)
            category_weight = category_weights.get(category, 0.3)
            
            # Step 2: Get impact
            impact = risk.details.get("impact_count", 0)
            
            # Step 3: Error probability modifier (for Inconsistent Formula only)
            error_probability = 1.0  # Default
            
            if risk.risk_type in ["Inconsistent Formula", "inconsistent_formula"]:
                likelihood = risk.details.get("likelihood_assessment", "")
                
                if "Likely error" in likelihood:
                    error_probability = 1.0  # 100%
                elif "Uncertain" in likelihood:
                    error_probability = 0.8  # 80%
                elif "Possibly intentional" in likelihood:
                    error_probability = 0.5  # 50%
                elif "Likely intentional" in likelihood:
                    error_probability = 0.3  # 30%
            
            # Step 4: Calculate risk score
            risk_score = category_weight * impact * error_probability
            
            # Step 5: Determine severity
            if risk_score >= 50:
                new_severity = "Critical"
            elif risk_score >= 20:
                new_severity = "High"
            elif risk_score >= 5:
                new_severity = "Medium"
            else:
                new_severity = "Low"
            
            # Update risk
            risk.severity = new_severity
            risk.details["risk_score"] = round(risk_score, 1)
            risk.details["category_weight"] = category_weight
            risk.details["error_probability"] = error_probability
        
        return risks
    
    def _get_context_labels(self, sheet: str, cell_address: str, 
                           cells: Dict[str, CellInfo], label_columns: str = "A:D") -> tuple:
        """
        Find row and column labels for a cell.
        
        DIAGNOSTIC FEATURE 1: Multi-Column Context Selector
        Supports concatenating labels from multiple columns to capture hierarchy.
        
        Args:
            sheet: Sheet name
            cell_address: Cell address (e.g., "E92")
            cells: Dictionary of all cells
            label_columns: Column range for labels (e.g., "A:D" or "B:C")
            
        Returns:
            Tuple of (row_label, col_label)
        """
        import re
        from openpyxl.utils import column_index_from_string, get_column_letter
        
        # Parse cell address (e.g., "E92" -> col="E", row=92)
        match = re.match(r'^([A-Z]+)(\d+)$', cell_address)
        if not match:
            return None, None
        
        col_letter = match.group(1)
        row_num = int(match.group(2))
        col_num = column_index_from_string(col_letter)
        
        # Find row label: Get the BEST text label to the left
        # Strategy: Scan ALL cells from target to A column, then pick the best candidate
        # Priority: Leftmost non-annotation text (avoid units like "(千円)" or notes like "※")
        row_label = None
        candidates = []  # Store all text labels found
        
        try:
            # Scan leftward from target cell to column A (collect ALL candidates)
            for check_col in range(col_num - 1, 0, -1):
                check_col_letter = get_column_letter(check_col)
                key = f"{sheet}!{check_col_letter}{row_num}"
                cell = cells.get(key)
                
                if not cell:
                    continue  # Skip if cell doesn't exist, but keep scanning
                
                # Get cell value (even if it's a formula cell)
                value = cell.value
                
                if not value:
                    continue  # Skip empty cells, but keep scanning
                
                # CRITICAL FIX: Check if VALUE is text (not if cell HAS formula)
                # Many cells have formulas but display text labels
                # We want to accept text labels regardless of whether they're from formulas
                
                # Only accept string values (text labels)
                if isinstance(value, str):
                    # NUCLEAR TRIM: Handle Japanese full-width spaces
                    value_str = value.replace('\u3000', ' ').strip()
                    
                    # Skip formula strings (e.g., "=A1+B1" displayed as text)
                    if value_str and not value_str.startswith('='):
                        # Store candidate with column position
                        candidates.append({
                            'text': value_str,
                            'col': check_col,
                            'col_letter': check_col_letter
                        })
            
            # Select BEST candidate (prefer leftmost, non-annotation text)
            if candidates:
                # Score each candidate
                scored_candidates = []
                for candidate in candidates:
                    text = candidate['text']
                    col = candidate['col']
                    
                    score = 0
                    
                    # Priority 1: Column position (item names are usually in first 20-30 columns)
                    # Based on real-world Excel usage patterns:
                    # - Columns A-T (1-20): Very likely item names → High score
                    # - Columns U-AD (21-30): Possibly item names → Medium score
                    # - Columns AE+ (31+): Unlikely item names → Low/negative score
                    
                    if col <= 20:  # A-T columns (1-20)
                        # High score: 300 (A) down to 100 (T)
                        score += 300 - (col - 1) * 10
                    elif col <= 30:  # U-AD columns (21-30)
                        # Medium score: 90 down to 0
                        score += 90 - (col - 21) * 9
                    else:  # AE+ columns (31+)
                        # Low/negative score: -10, -20, -30...
                        score -= (col - 30) * 10
                    
                    # Priority 2: Avoid annotations (parentheses, symbols)
                    if '(' in text or ')' in text or '（' in text or '）' in text:
                        score -= 50  # Heavy penalty for parentheses (units/notes)
                    
                    if text.startswith('※') or text.startswith('*') or text.startswith('注'):
                        score -= 50  # Heavy penalty for note markers
                    
                    # Priority 3: Penalize very short text (units are usually 1-2 chars)
                    # Note: Long text could be annotations, so we don't give bonus for length
                    if len(text) <= 2:
                        score -= 20  # Penalty for very short text (likely units like "円", "%")
                    
                    # Priority 4: Avoid pure symbols
                    if text in ['-', '=', '/', '\\', '|', '・', '…']:
                        score -= 100
                    
                    scored_candidates.append({
                        'text': text,
                        'col': col,
                        'col_letter': candidate['col_letter'],
                        'score': score
                    })
                
                # Pick candidate with highest score
                best_candidate = max(scored_candidates, key=lambda x: x['score'])
                row_label = best_candidate['text']
            
        except Exception as e:
            # Fallback: Simple leftward scan if scoring logic fails
            print(f"[Context] Error in smart label selection: {e}. Using fallback.")
            row_label = None
            
            for check_col in range(col_num - 1, 0, -1):
                check_col_letter = get_column_letter(check_col)
                key = f"{sheet}!{check_col_letter}{row_num}"
                cell = cells.get(key)
                
                if not cell or not cell.value:
                    continue
                
                if cell.formula:
                    continue
                
                value = cell.value
                
                if isinstance(value, str):
                    value_str = value.replace('\u3000', ' ').strip()
                    
                    if value_str and not value_str.startswith('='):
                        row_label = value_str
                        break
        
        # Strategy 2: If nothing found on left, check row above (header)
        if not row_label and row_num > 1:
            key = f"{sheet}!{col_letter}{row_num - 1}"
            cell = cells.get(key)
            
            if cell:
                # Apply same TYPE FILTER
                if cell.formula:
                    # Formula cell - reject
                    pass
                elif cell.value:
                    value = cell.value
                    
                    if isinstance(value, str):
                        # NUCLEAR TRIM: Handle Japanese full-width spaces (\u3000) and all whitespace
                        value_str = value.replace('\u3000', ' ').strip()
                        
                        # REJECT empty strings after nuclear trim
                        if not value_str:
                            pass  # Treat as None
                        # REJECT formulas
                        elif value_str.startswith('='):
                            pass  # Treat as None
                        # ACCEPT non-empty text
                        else:
                            row_label = value_str
        
        # PHASE 5: AI recovery moved to _add_context_labels() for better control
        
        # Find column label: Scan rows 1-20 in the same column
        col_label = None
        for check_row in range(1, min(21, row_num)):  # Rows 1-20
            key = f"{sheet}!{col_letter}{check_row}"
            cell = cells.get(key)
            
            # FIX 1: "Dirty Header" Bug - Check for formulas in column headers
            if cell and cell.value:
                # CRITICAL: Reject formula cells (checksum formulas in headers)
                if cell.formula:
                    continue  # Skip formula cells
                
                # Handle datetime objects - convert to simple date format
                from datetime import datetime
                if isinstance(cell.value, datetime):
                    # Format as YYYY-MM for column headers
                    value_str = cell.value.strftime('%Y-%m')
                else:
                    # NUCLEAR TRIM: Handle Japanese full-width spaces (\u3000) and all whitespace
                    value_str = str(cell.value).replace('\u3000', ' ').strip()
                
                # REJECT empty strings after nuclear trim
                if not value_str:
                    continue  # Skip whitespace-only cells
                
                # Double-check: Reject if starts with =
                if value_str.startswith('='):
                    continue  # Skip formula strings
                
                # Check if it matches a date pattern or contains FY/Q
                # Date patterns: "04-2024", "2024-04", "Apr 2024", "Q1 2024", "FY2024"
                date_patterns = [
                    r'\d{2}-\d{4}',  # 04-2024
                    r'\d{4}-\d{2}',  # 2024-04
                    r'[A-Z][a-z]{2}\s+\d{4}',  # Apr 2024
                    r'Q\d',  # Q1, Q2, etc.
                    r'FY\s*\d{4}',  # FY2024, FY 2024
                ]
                
                for pattern in date_patterns:
                    if re.search(pattern, value_str):
                        col_label = value_str
                        break
                
                if col_label:
                    break
        
        return row_label, col_label
    
    def _calculate_health_score(self, risks: List[RiskAlert]) -> int:
        """
        Calculate health score with category-based weighting.
        
        Category Weights (relative to Fatal Error):
        - Fatal Error: 100% (base)
        - Integrity Risk: 50% (half of Fatal)
        - Structural Debt: 10% (one-tenth of Fatal)
        
        Severity Penalties (Fatal Error base):
        - Critical: -5 points
        - High: -4 points
        - Medium: -3 points
        - Low: -1 point
        
        Floor: Minimum 30 (psychological safety)
        
        Args:
            risks: List of detected risks
            
        Returns:
            Health score (30-100)
        """
        from src.models import RiskCategory
        
        score = 100.0
        
        # Define base penalties for Fatal Error
        fatal_penalties = {
            "Critical": 5.0,
            "High": 4.0,
            "Medium": 3.0,
            "Low": 1.0
        }
        
        # Category multipliers
        category_multipliers = {
            RiskCategory.FATAL_ERROR: 1.0,      # 100%
            RiskCategory.INTEGRITY_RISK: 0.5,   # 50%
            RiskCategory.STRUCTURAL_DEBT: 0.1   # 10%
        }
        
        # Classify risks first (if not already classified)
        triage = RiskTriageEngine(risks)
        triage.classify_all()
        
        # Apply penalties for each risk
        for risk in risks:
            category = risk.category
            severity = risk.severity
            
            # Skip if category or severity is not recognized
            if category not in category_multipliers:
                continue
            
            if severity not in fatal_penalties:
                continue
            
            # Calculate penalty
            base_penalty = fatal_penalties[severity]
            multiplier = category_multipliers[category]
            penalty = base_penalty * multiplier
            
            score -= penalty
        
        # Floor: Minimum 30 (psychological safety)
        return max(30, int(score))
    
    def trace_to_drivers(self, model: ModelAnalysis, cell_address: str) -> List[str]:
        """
        Trace a cell to all ultimate drivers (cells with no outgoing edges).
        
        This is critical for understanding the impact of hardcoded values.
        For example, if a hardcoded cell affects "Revenue", "EBITDA", and "Net Income",
        this method will return all three driver cells.
        
        Args:
            model: ModelAnalysis object with dependency graph
            cell_address: Starting cell address in "Sheet!Address" format
            
        Returns:
            List of driver cell addresses (cells with no dependents)
            
        Note: This method handles Virtual Fill cells correctly. If a driver is
        inside a merged range, all virtual cells in that range are considered.
        """
        if cell_address not in model.dependency_graph:
            return []
        
        # Find all reachable nodes from this cell
        try:
            # Use BFS to find all descendants
            descendants = nx.descendants(model.dependency_graph, cell_address)
        except nx.NetworkXError:
            return []
        
        # Filter to only ultimate drivers (nodes with no outgoing edges)
        drivers = []
        for node in descendants:
            # Check if this node has any dependents
            if model.dependency_graph.out_degree(node) == 0:
                drivers.append(node)
        
        # Also check if the starting cell itself is a driver
        if model.dependency_graph.out_degree(cell_address) == 0:
            drivers.append(cell_address)
        
        return drivers

    # ========================================================================
    # Excel Rehab Maturity Scoring (Phase 7)
    # ========================================================================
    
    def calculate_maturity_level_heuristic(self, model: ModelAnalysis) -> 'MaturityScore':
        """
        Fast heuristic-based maturity scoring for initial diagnosis.
        
        Target: Complete within 3 seconds of file upload.
        
        Heuristic Rules:
        1. Critical risks present → Level 1 or 2 (depending on type)
        2. Hardcode count > 5 → Level 1 (Static Model)
        3. High risks > 3 → Level 2 (Unstable Model)
        4. Clean model → Level 3 (Strategic Model)
        
        Args:
            model: ModelAnalysis object with risks detected
            
        Returns:
            MaturityScore with level and counts
        """
        from src.models import MaturityLevel, MaturityScore
        import time
        
        start_time = time.time()
        
        # Get risk counts
        risk_counts = model.get_risk_counts()
        critical_count = risk_counts["Critical"]
        high_count = risk_counts["High"]
        
        # Count hardcodes
        hardcode_count = len([r for r in model.risks if r.risk_type == "Hidden Hardcode"])
        
        # Heuristic 1: Critical risks present
        if critical_count > 0:
            # Check if circular references exist
            has_circular = any(r.risk_type == "Circular Reference" for r in model.risks)
            if has_circular:
                level = MaturityLevel.LEVEL_2  # Unstable
            else:
                level = MaturityLevel.LEVEL_1  # Static (likely hardcodes)
        
        # Heuristic 2: High hardcode count → Level 1
        elif hardcode_count > 5:
            level = MaturityLevel.LEVEL_1  # Static
        
        # Heuristic 3: Some high risks → Level 2
        elif high_count > 3:
            level = MaturityLevel.LEVEL_2  # Unstable
        
        # Heuristic 4: Clean model → Level 3
        else:
            level = MaturityLevel.LEVEL_3  # Healthy
        
        # Calculate progress to next level
        progress = self._calculate_progress_to_next_level(
            level, hardcode_count, critical_count, high_count
        )
        
        elapsed = time.time() - start_time
        print(f"[Maturity] Heuristic scoring completed in {elapsed:.2f}s")
        
        return MaturityScore(
            level=level,
            hardcode_count=hardcode_count,
            critical_count=critical_count,
            high_count=high_count,
            progress_to_next=progress
        )
    
    def calculate_maturity_level_deep(self, model: ModelAnalysis) -> 'MaturityScore':
        """
        Accurate maturity scoring after full dependency analysis.
        
        This method performs deeper analysis to identify critical rows
        and calculate more accurate maturity levels.
        
        Deep Analysis Rules:
        1. Level 1: > 5 hardcodes in critical rows (KPI-related, high impact)
        2. Level 2: Circular refs OR > 3 high-severity risks
        3. Level 3: No Critical risks AND < 3 High risks
        
        Args:
            model: ModelAnalysis object with full dependency graph
            
        Returns:
            MaturityScore with accurate level and counts
        """
        from src.models import MaturityLevel, MaturityScore
        
        # Get risk counts
        risk_counts = model.get_risk_counts()
        critical_count = risk_counts["Critical"]
        high_count = risk_counts["High"]
        
        # Count hardcodes
        hardcode_risks = [r for r in model.risks if r.risk_type == "Hidden Hardcode"]
        hardcode_count = len(hardcode_risks)
        
        # Identify critical rows (simplified: use all hardcodes for MVP)
        # In future: Use AI or heuristics to identify KPI rows
        critical_hardcodes = hardcode_count
        
        # Level 1: High hardcode count in critical rows
        if critical_hardcodes > 5:
            level = MaturityLevel.LEVEL_1
        
        # Level 2: Circular refs or high-severity risks
        elif critical_count > 0 or high_count > 3:
            level = MaturityLevel.LEVEL_2
        
        # Level 3: Clean model
        else:
            level = MaturityLevel.LEVEL_3
        
        # Calculate progress to next level
        progress = self._calculate_progress_to_next_level(
            level, hardcode_count, critical_count, high_count
        )
        
        return MaturityScore(
            level=level,
            hardcode_count=hardcode_count,
            critical_count=critical_count,
            high_count=high_count,
            progress_to_next=progress
        )
    
    def _calculate_progress_to_next_level(
        self, 
        level: 'MaturityLevel', 
        hardcode_count: int, 
        critical_count: int, 
        high_count: int
    ) -> float:
        """
        Calculate progress percentage toward next maturity level.
        
        Args:
            level: Current maturity level
            hardcode_count: Number of hardcodes
            critical_count: Number of critical risks
            high_count: Number of high-severity risks
            
        Returns:
            Progress percentage (0-100)
        """
        from src.models import MaturityLevel
        
        if level == MaturityLevel.LEVEL_1:
            # Progress to Level 2: Need to reduce hardcodes to ≤ 5
            if hardcode_count > 5:
                target = 5
                current = hardcode_count
                fixed = max(0, current - target)
                total_to_fix = current - target
                progress = (fixed / total_to_fix) * 100 if total_to_fix > 0 else 0
                return 100 - progress  # Invert: 0% when many issues, 100% when close
            else:
                return 100.0  # Ready to level up
        
        elif level == MaturityLevel.LEVEL_2:
            # Progress to Level 3: Need to eliminate critical risks and reduce high risks to < 3
            total_issues = critical_count + max(0, high_count - 3)
            if total_issues == 0:
                return 100.0  # Ready to level up
            
            # Assume starting point (for progress calculation)
            # This is simplified - in production, track historical data
            assumed_start = total_issues + 5
            fixed = assumed_start - total_issues
            progress = (fixed / assumed_start) * 100
            return min(100.0, max(0.0, progress))
        
        else:  # Level 3
            return 100.0  # Already at max level
    
    def calculate_unlock_requirements(self, maturity_score: 'MaturityScore') -> 'UnlockRequirement':
        """
        Calculate requirements to unlock the next maturity level.
        
        Args:
            maturity_score: Current maturity score
            
        Returns:
            UnlockRequirement with actionable steps
        """
        from src.models import MaturityLevel, UnlockRequirement
        
        level = maturity_score.level
        hardcode_count = maturity_score.hardcode_count
        critical_count = maturity_score.critical_count
        high_count = maturity_score.high_count
        
        if level == MaturityLevel.LEVEL_1:
            # To reach Level 2: Reduce hardcodes to ≤ 5
            hardcodes_to_fix = max(0, hardcode_count - 5)
            
            actionable_steps = []
            if hardcodes_to_fix > 0:
                actionable_steps.append(f"Replace {hardcodes_to_fix} hardcoded value(s) with cell references")
                actionable_steps.append("Click '✨ Suggest Improvement' on the Top 3 Killers for AI guidance")
            
            return UnlockRequirement(
                current_level=MaturityLevel.LEVEL_1,
                next_level=MaturityLevel.LEVEL_2,
                hardcodes_to_fix=hardcodes_to_fix,
                critical_risks_to_fix=0,
                high_risks_to_fix=0,
                actionable_steps=actionable_steps,
                progress_percentage=maturity_score.progress_to_next
            )
        
        elif level == MaturityLevel.LEVEL_2:
            # To reach Level 3: Eliminate critical risks and reduce high risks to < 3
            critical_to_fix = critical_count
            high_to_fix = max(0, high_count - 3)
            
            actionable_steps = []
            if critical_to_fix > 0:
                actionable_steps.append(f"Fix {critical_to_fix} critical risk(s) (circular references, etc.)")
            if high_to_fix > 0:
                actionable_steps.append(f"Fix {high_to_fix} high-severity risk(s)")
            if actionable_steps:
                actionable_steps.append("Use AI suggestions to identify and fix structural issues")
            
            return UnlockRequirement(
                current_level=MaturityLevel.LEVEL_2,
                next_level=MaturityLevel.LEVEL_3,
                hardcodes_to_fix=0,
                critical_risks_to_fix=critical_to_fix,
                high_risks_to_fix=high_to_fix,
                actionable_steps=actionable_steps,
                progress_percentage=maturity_score.progress_to_next
            )
        
        else:  # Level 3
            # Already at max level
            return UnlockRequirement(
                current_level=MaturityLevel.LEVEL_3,
                next_level=None,
                hardcodes_to_fix=0,
                critical_risks_to_fix=0,
                high_risks_to_fix=0,
                actionable_steps=["🎉 Congratulations! All features unlocked. Your model is ready for strategic planning."],
                progress_percentage=100.0
            )

    # ========================================================================
    # Impact Scoring System (Phase 7.2)
    # ========================================================================
    
    def calculate_impact_score(self, model: ModelAnalysis, cell_address: str, 
                               hardcoded_value: str) -> Dict[str, any]:
        """
        Calculate impact metrics for a hardcoded value.
        
        Metrics:
        - Diffusion: How many times does this value appear?
        - Dominance: How many dependent cells does this cell have?
        - Volatility: Does this row contain varying hardcodes?
        
        Args:
            model: ModelAnalysis object
            cell_address: Cell address (e.g., "Sheet1!A1")
            hardcoded_value: The hardcoded value to analyze
            
        Returns:
            Dict with diffusion, dominance, volatility, and prescription_mode
        """
        # Metric 1: Diffusion - Count occurrences across workbook
        diffusion = self._calculate_diffusion(model, hardcoded_value)
        
        # Metric 2: Dominance - Count all descendants in dependency graph
        dominance = self._calculate_dominance(model, cell_address)
        
        # Metric 3: Volatility - Check for varying hardcodes in same row
        volatility = self._calculate_volatility(model, cell_address)
        
        # Determine prescription mode based on metrics
        prescription_mode = self._determine_prescription_mode(diffusion, dominance, volatility)
        
        return {
            'diffusion': diffusion,
            'dominance': dominance,
            'volatility': volatility,
            'prescription_mode': prescription_mode
        }
    
    def _calculate_diffusion(self, model: ModelAnalysis, hardcoded_value: str) -> int:
        """
        Count how many times this value appears in the workbook.
        
        Args:
            model: ModelAnalysis object
            hardcoded_value: The value to count
            
        Returns:
            Number of occurrences
        """
        count = 0
        for risk in model.risks:
            if risk.risk_type == "Hidden Hardcode":
                if risk.details.get('hardcoded_value') == hardcoded_value:
                    count += 1
        return count
    
    def _calculate_dominance(self, model: ModelAnalysis, cell_address: str) -> int:
        """
        Count all dependent cells (children + grandchildren + ...).
        
        Uses NetworkX to traverse the dependency graph and count all descendants.
        
        Args:
            model: ModelAnalysis object
            cell_address: Cell address in "Sheet!Address" format
            
        Returns:
            Number of dependent cells
        """
        if cell_address not in model.dependency_graph:
            return 0
        
        try:
            # Get all descendants (children, grandchildren, etc.)
            descendants = nx.descendants(model.dependency_graph, cell_address)
            return len(descendants)
        except:
            return 0
    
    def _calculate_volatility(self, model: ModelAnalysis, cell_address: str) -> str:
        """
        Check if this row contains varying hardcodes.
        
        Volatility is HIGH if the same row has multiple different hardcoded values,
        suggesting scenario-based logic.
        
        Args:
            model: ModelAnalysis object
            cell_address: Cell address in "Sheet!Address" format
            
        Returns:
            "High" or "Low"
        """
        try:
            # Extract sheet and row number
            sheet, address = cell_address.split('!')
            row_match = re.match(r'[A-Z]+(\d+)', address)
            if not row_match:
                return "Low"
            
            row_num = int(row_match.group(1))
            
            # Find all hardcodes in the same row
            hardcodes_in_row = set()
            for risk in model.risks:
                if risk.risk_type == "Hidden Hardcode" and risk.sheet == sheet:
                    risk_row_match = re.match(r'[A-Z]+(\d+)', risk.cell)
                    if risk_row_match and int(risk_row_match.group(1)) == row_num:
                        hardcodes_in_row.add(risk.details.get('hardcoded_value', ''))
            
            # High volatility if 3+ different values in same row
            return "High" if len(hardcodes_in_row) >= 3 else "Low"
        except:
            return "Low"
    
    def _determine_prescription_mode(self, diffusion: int, dominance: int, 
                                     volatility: str) -> str:
        """
        Determine the prescription mode based on impact metrics.
        
        Rules:
        - Diffusion > 3: "Centralization" (value scattered across workbook)
        - Dominance > 50: "Driver Decomposition" (high impact on other cells)
        - Volatility High: "Scenario Planning" (multiple values suggest scenarios)
        
        Args:
            diffusion: Number of occurrences
            dominance: Number of dependent cells
            volatility: "High" or "Low"
            
        Returns:
            Prescription mode string
        """
        # Priority order: Volatility > Dominance > Diffusion
        if volatility == "High":
            return "Scenario Planning"
        elif dominance > 50:
            return "Driver Decomposition"
        elif diffusion > 3:
            return "Centralization"
        else:
            return "Basic Refactoring"

    # ========================================================================
    # DIAGNOSTIC SUITE - 5 Advanced Logic Checks (December 2025)
    # ========================================================================
    
    def _detect_row_inconsistency(self, cells: Dict[str, CellInfo]) -> List[RiskAlert]:
        """
        DIAGNOSTIC FEATURE 3: Row Consistency Scanner (Horizontal Check)
        
        Detects formulas that break the pattern in a row - catches drag-and-drop errors.
        
        Logic:
        - Group cells by row
        - Convert formulas to R1C1 notation (relative pattern)
        - If 10 cells have Pattern A but 1 cell has Pattern B (or hardcode), flag it
        
        Example:
        Row 5: =B5*C5, =B5*C5, =B5*C5, =B5*D5 (INCONSISTENT!)
        R1C1:  =RC[-3]*RC[-2], =RC[-3]*RC[-2], =RC[-3]*RC[-2], =RC[-3]*RC[-1]
        
        Value: Catches copy-paste errors where one cell has wrong formula
        
        Args:
            cells: Dictionary of CellInfo objects
            
        Returns:
            List of RiskAlert objects for inconsistent formulas
        """
        risks = []
        
        # Group cells by sheet and row
        from collections import defaultdict
        rows = defaultdict(lambda: defaultdict(list))  # {sheet: {row_num: [cells]}}
        
        for cell_addr, cell in cells.items():
            if cell.formula:
                sheet, address = cell_addr.split('!')
                # Extract row number
                match = re.match(r'([A-Z]+)(\d+)', address)
                if match:
                    row_num = int(match.group(2))
                    rows[sheet][row_num].append((address, cell))
        
        # Check each row for consistency
        for sheet, sheet_rows in rows.items():
            for row_num, row_cells in sheet_rows.items():
                if len(row_cells) < 3:  # Need at least 3 cells to detect pattern
                    continue
                
                # Convert formulas to R1C1 patterns
                patterns = {}
                for address, cell in row_cells:
                    pattern = self._formula_to_r1c1_pattern(address, cell.formula)
                    if pattern not in patterns:
                        patterns[pattern] = []
                    patterns[pattern].append((address, cell))
                
                # Check if one pattern dominates
                if len(patterns) > 1:
                    pattern_counts = {p: len(cells) for p, cells in patterns.items()}
                    max_count = max(pattern_counts.values())
                    min_count = min(pattern_counts.values())
                    total_cells = len(row_cells)
                    
                    # If majority pattern exists (>= 70% of cells)
                    if max_count >= total_cells * 0.7:
                        # Find the dominant pattern
                        dominant_pattern = max(patterns.keys(), key=lambda p: len(patterns[p]))
                        
                        # Calculate majority percentage for assessment
                        majority_percentage = (max_count / total_cells) * 100
                        
                        # Flag cells that don't match dominant pattern
                        for pattern, pattern_cells in patterns.items():
                            if pattern != dominant_pattern:
                                minority_count = len(pattern_cells)
                                
                                # Calculate percentage of inconsistent cells
                                inconsistent_percentage = (minority_count / total_cells) * 100
                                
                                # Assess likelihood of intentional vs error
                                if minority_count == 1:
                                    # Single cell difference - likely intentional
                                    severity = "Low"
                                    likelihood_assessment = "意図的な変更の可能性が高い"
                                    description = f"この行の他の{max_count}個のセルとは1箇所だけ数式パターンが異なります。意図的な変更の可能性が高いですが、念のため確認してください。"
                                else:
                                    # Multiple cells different - likely error
                                    severity = "High"
                                    likelihood_assessment = "エラーの可能性が高い"
                                    description = f"この行の他の{max_count}個のセルと、{minority_count}セル（{inconsistent_percentage:.0f}%）数式パターンが異なります。確認してください。"
                                
                                for address, cell in pattern_cells:
                                    risks.append(RiskAlert(
                                        risk_type="Inconsistent Formula",
                                        severity=severity,
                                        sheet=sheet,
                                        cell=address,
                                        description=description,
                                        details={
                                            'formula': cell.formula,
                                            'pattern': pattern,
                                            'dominant_pattern': dominant_pattern,
                                            'row': row_num,
                                            'consistent_count': max_count,
                                            'minority_count': minority_count,
                                            'majority_percentage': majority_percentage,
                                            'likelihood_assessment': likelihood_assessment
                                        }
                                    ))
        
        return risks
    
    def _formula_to_r1c1_pattern(self, cell_address: str, formula: str) -> str:
        """
        Convert formula to R1C1 pattern for consistency checking.
        
        Example:
        Cell E5: =B5*C5 -> =RC[-3]*RC[-2]
        Cell F5: =C5*D5 -> =RC[-3]*RC[-2] (same pattern!)
        Cell G5: =D5*E5 -> =RC[-3]*RC[-2] (same pattern!)
        Cell H5: =E5*F5 -> =RC[-3]*RC[-2] (same pattern!)
        Cell I5: =F5*H5 -> =RC[-3]*RC[-1] (DIFFERENT pattern!)
        
        Args:
            cell_address: Cell address (e.g., "E5")
            formula: Formula string
            
        Returns:
            R1C1 pattern string
        """
        from openpyxl.utils import column_index_from_string
        
        # Extract current cell position
        match = re.match(r'([A-Z]+)(\d+)', cell_address)
        if not match:
            return formula
        
        curr_col = column_index_from_string(match.group(1))
        curr_row = int(match.group(2))
        
        # Replace all cell references with R1C1 notation
        pattern = formula
        
        # Find all cell references (e.g., A1, B5, AA10)
        cell_refs = re.findall(r'\b([A-Z]+)(\d+)\b', formula)
        
        for col_letter, row_str in cell_refs:
            ref_col = column_index_from_string(col_letter)
            ref_row = int(row_str)
            
            # Calculate relative offsets
            row_offset = ref_row - curr_row
            col_offset = ref_col - curr_col
            
            # Build R1C1 notation
            if row_offset == 0 and col_offset == 0:
                r1c1 = "RC"
            elif row_offset == 0:
                r1c1 = f"RC[{col_offset}]"
            elif col_offset == 0:
                r1c1 = f"R[{row_offset}]C"
            else:
                r1c1 = f"R[{row_offset}]C[{col_offset}]"
            
            # Replace in pattern
            pattern = pattern.replace(f"{col_letter}{row_str}", r1c1, 1)
        
        return pattern
    
    def _detect_value_conflicts(self, cells: Dict[str, CellInfo]) -> List[RiskAlert]:
        """
        DIAGNOSTIC FEATURE 4: Value Consistency Guard (Update Omission)
        
        Detects hardcoded values that should be the same but aren't - catches update omissions.
        
        Logic:
        - Group hardcoded cells by their row/column label
        - If label "Tax Rate" has values {0.3} in 10 cells but {0.35} in 1 cell, flag it
        
        Example:
        Tax Rate row: 0.3, 0.3, 0.3, 0.35, 0.3 <- One cell not updated!
        
        Value: Catches when user updates most cells but forgets one
        
        Args:
            cells: Dictionary of CellInfo objects
            
        Returns:
            List of RiskAlert objects for conflicting values
        """
        risks = []
        
        # First, we need to get labels for cells (this requires context)
        # Group hardcoded values by their label
        from collections import defaultdict
        label_values = defaultdict(lambda: defaultdict(list))  # {label: {value: [cells]}}
        
        for cell_addr, cell in cells.items():
            # Check if cell has hardcoded value (no formula, numeric value)
            if not cell.formula and isinstance(cell.value, (int, float)):
                sheet, address = cell_addr.split('!')
                
                # Get row label for this cell
                row_label, col_label = self._get_context_labels(sheet, address, cells)
                
                if row_label:  # Only check if we have a label
                    # Normalize value to 2 decimal places for comparison
                    normalized_value = round(float(cell.value), 2)
                    label_values[row_label][normalized_value].append((sheet, address, cell))
        
        # Check each label for conflicting values
        for label, values_dict in label_values.items():
            if len(values_dict) > 1:  # Multiple different values for same label
                # Find the dominant value (most common)
                value_counts = {v: len(cells) for v, cells in values_dict.items()}
                max_count = max(value_counts.values())
                total_count = sum(value_counts.values())
                
                # If dominant value exists (>= 70% of cells)
                if max_count >= total_count * 0.7 and total_count >= 3:
                    dominant_value = max(values_dict.keys(), key=lambda v: len(values_dict[v]))
                    
                    # Flag cells with non-dominant values
                    for value, value_cells in values_dict.items():
                        if value != dominant_value:
                            for sheet, address, cell in value_cells:
                                risks.append(RiskAlert(
                                    risk_type="Conflicting Value",
                                    severity="High",
                                    sheet=sheet,
                                    cell=address,
                                    description=f"Value {value} differs from {max_count} other cells with label '{label}' (expected {dominant_value})",
                                    details={
                                        'value': value,
                                        'expected_value': dominant_value,
                                        'label': label,
                                        'consistent_count': max_count,
                                        'total_count': total_count
                                    }
                                ))
        
        return risks
    
    def _detect_external_links(self, cells: Dict[str, CellInfo]) -> List[RiskAlert]:
        """
        DIAGNOSTIC FEATURE 5: Phantom Link Detector (External References)
        
        Detects external file references that will break when file is shared.
        
        Logic:
        - Scan formulas for brackets [ ] which indicate external workbook references
        - Internal cross-sheet references (Sheet2!A1) are NOT flagged
        - Only flag true external workbook links: =[File.xlsx]Sheet!A1
        
        Example:
        ='[Budget2024.xlsx]Sheet1'!A5  ← External (flagged)
        =Sheet2!A1  ← Internal (NOT flagged)
        ='Sheet Name'!A1  ← Internal (NOT flagged)
        
        Value: Prevents broken links when sharing files
        
        Args:
            cells: Dictionary of CellInfo objects
            
        Returns:
            List of RiskAlert objects for external links
        """
        risks = []
        
        for cell_addr, cell in cells.items():
            if cell.formula:
                sheet, address = cell_addr.split('!')
                
                # Check for external workbook indicators
                # Brackets [ ] specifically indicate external workbook references in Excel
                has_brackets = '[' in cell.formula and ']' in cell.formula
                
                # Only flag if brackets are present (true external workbook reference)
                # Do NOT flag internal cross-sheet references like =Sheet2!A1
                if has_brackets:
                    # Extract the external file name
                    external_file = "Unknown"
                    bracket_match = re.search(r'\[([^\]]+)\]', cell.formula)
                    if bracket_match:
                        external_file = bracket_match.group(1)
                    
                    risks.append(RiskAlert(
                        risk_type="External Link",
                        severity="Medium",
                        sheet=sheet,
                        cell=address,
                        description=f"Formula references external file: {external_file}",
                        details={
                            'formula': cell.formula,
                            'external_file': external_file
                        }
                    ))
        
        return risks
    
    def _detect_formula_errors(self, cells: Dict[str, CellInfo]) -> List[RiskAlert]:
        """
        Detect Excel formula errors (#REF!, #DIV/0!, #VALUE!, #NAME?, #N/A, #NUM!, #NULL!).
        
        These errors indicate broken formulas that prevent the model from calculating correctly.
        All formula errors are classified as FATAL (Critical severity).
        
        Error Types:
        - #REF!: Reference to deleted cell/sheet
        - #DIV/0!: Division by zero
        - #VALUE!: Wrong type of argument
        - #NAME?: Unrecognized function/name
        - #N/A: Value not available
        - #NUM!: Invalid numeric value
        - #NULL!: Incorrect range operator
        
        Args:
            cells: Dictionary of CellInfo objects
            
        Returns:
            List of RiskAlert objects for formula errors
        """
        risks = []
        
        # Excel error patterns
        error_patterns = {
            '#REF!': 'Reference to deleted cell or sheet',
            '#DIV/0!': 'Division by zero',
            '#VALUE!': 'Wrong type of argument or operand',
            '#NAME?': 'Unrecognized function or name',
            '#N/A': 'Value not available',
            '#NUM!': 'Invalid numeric value',
            '#NULL!': 'Incorrect range operator'
        }
        
        for cell_addr, cell in cells.items():
            # Check if cell value contains an error
            if cell.value and isinstance(cell.value, str):
                for error_code, error_desc in error_patterns.items():
                    if error_code in str(cell.value):
                        sheet, address = cell_addr.split('!')
                        
                        risks.append(RiskAlert(
                            risk_type="Formula Error",
                            severity="Critical",  # All formula errors are FATAL
                            sheet=sheet,
                            cell=address,
                            description=f"{error_code}: {error_desc}",
                            details={
                                'error_code': error_code,
                                'error_description': error_desc,
                                'formula': cell.formula or '',
                                'value': cell.value
                            }
                        ))
                        break  # Only report first error found in cell
        
        return risks
    
    def translate_formula_to_labels(self, formula: str, cell_address: str, 
                                    cells: Dict[str, CellInfo], sheet: str) -> str:
        """
        DIAGNOSTIC FEATURE 2: Logic Translator (Vertical Check)
        
        Translates formula from cell references to semantic labels.
        Makes semantic errors obvious (e.g., Price + Cost instead of Price * Quantity).
        
        Example:
        Formula: =F12*F13
        Translated: =[Unit Price] * [Quantity]
        
        Semantic Error Example:
        Formula: =F12+F14
        Translated: =[Unit Price] + [Cost]  <- Addition instead of multiplication!
        
        Args:
            formula: Formula string (e.g., "=F12*F13")
            cell_address: Current cell address (e.g., "G12")
            cells: Dictionary of all cells
            sheet: Sheet name
            
        Returns:
            Translated formula with labels
        """
        if not formula:
            return ""
        
        translated = formula
        
        # Pattern 1: Cross-sheet references with quotes: 'Sheet Name'!A1
        cross_sheet_pattern = r"'([^']+)'!([A-Z]+\d+)"
        cross_sheet_refs = re.findall(cross_sheet_pattern, formula)
        
        for ref_sheet, ref_cell in cross_sheet_refs:
            # Get label for this cross-sheet reference
            ref_full_addr = f"{ref_sheet}!{ref_cell}"
            
            if ref_full_addr in cells:
                row_label, col_label = self._get_context_labels(ref_sheet, ref_cell, cells)
                
                # Build label
                if row_label and col_label:
                    label = f"[{ref_sheet}:{row_label} @ {col_label}]"
                elif row_label:
                    label = f"[{ref_sheet}:{row_label}]"
                elif col_label:
                    label = f"[{ref_sheet}:{col_label}]"
                else:
                    label = f"[{ref_sheet}!{ref_cell}]"  # Fallback to full address
                
                # Replace in formula (preserve the original pattern)
                original_pattern = f"'{ref_sheet}'!{ref_cell}"
                translated = translated.replace(original_pattern, label, 1)
        
        # Pattern 2: Cross-sheet references without quotes: Sheet1!A1
        simple_cross_sheet_pattern = r"([A-Za-z0-9_]+)!([A-Z]+\d+)"
        simple_cross_refs = re.findall(simple_cross_sheet_pattern, formula)
        
        for ref_sheet, ref_cell in simple_cross_refs:
            # Skip if already processed (quoted version)
            if f"'{ref_sheet}'!{ref_cell}" in formula:
                continue
            
            # Get label for this cross-sheet reference
            ref_full_addr = f"{ref_sheet}!{ref_cell}"
            
            if ref_full_addr in cells:
                row_label, col_label = self._get_context_labels(ref_sheet, ref_cell, cells)
                
                # Build label
                if row_label and col_label:
                    label = f"[{ref_sheet}:{row_label} @ {col_label}]"
                elif row_label:
                    label = f"[{ref_sheet}:{row_label}]"
                elif col_label:
                    label = f"[{ref_sheet}:{col_label}]"
                else:
                    label = f"[{ref_sheet}!{ref_cell}]"  # Fallback to full address
                
                # Replace in formula
                original_pattern = f"{ref_sheet}!{ref_cell}"
                translated = translated.replace(original_pattern, label, 1)
        
        # Pattern 3: Same-sheet references: A1, B2, etc.
        same_sheet_refs = re.findall(r'\b([A-Z]+\d+)\b', formula)
        
        for ref in same_sheet_refs:
            # Skip if this looks like it was already processed as part of cross-sheet ref
            if f"!{ref}" in formula:
                continue
            
            # Get label for this same-sheet reference
            ref_full_addr = f"{sheet}!{ref}"
            
            if ref_full_addr in cells:
                row_label, col_label = self._get_context_labels(sheet, ref, cells)
                
                # Build label
                if row_label and col_label:
                    label = f"[{row_label} @ {col_label}]"
                elif row_label:
                    label = f"[{row_label}]"
                elif col_label:
                    label = f"[{col_label}]"
                else:
                    label = f"[{ref}]"  # Fallback to cell address
                
                # Replace in formula (only first occurrence to avoid issues)
                translated = translated.replace(ref, label, 1)
        
        return translated


# ============================================================================
# 3-Tier Risk Triage System (Phase 8)
# ============================================================================

def classify_risk(risk: RiskAlert, all_risks: List[RiskAlert] = None) -> RiskCategory:
    """
    Classify risk by business impact for 3-tier triage system.
    
    Classification Logic:
    - Fatal Errors (Tab 1): Model is broken or uncomputable
    - Integrity Risks (Tab 2): Model runs but logic/values seem wrong (HIGHEST PRIORITY)
    - Structural Debt (Tab 3): Works correctly but hard to maintain
    
    Args:
        risk: The risk to classify
        all_risks: All risks (needed for hardcode consistency check)
        
    Returns:
        RiskCategory enum value
    """
    
    # Tab 1: Fatal Errors (Calculation Breakage)
    if risk.risk_type in ["circular_reference", "Circular Reference"]:
        return RiskCategory.FATAL_ERROR
    
    if risk.risk_type in ["phantom_link", "Phantom Link", "external_link", "External Link"]:
        return RiskCategory.FATAL_ERROR
    
    if risk.risk_type in ["formula_error", "Formula Error"]:
        # #REF!, #VALUE!, #DIV/0!, etc.
        return RiskCategory.FATAL_ERROR
    
    # Tab 2: Integrity Risks (Suspicion of Error) - HIGHEST PRIORITY
    if risk.risk_type in ["inconsistent_formula", "Inconsistent Formula"]:
        # Row pattern breaks - logic may be wrong
        # BUT: If Impact = 0, it's Structural Debt (future risk, not active)
        impact_count = risk.details.get("impact_count", 0)
        if impact_count == 0:
            return RiskCategory.STRUCTURAL_DEBT
        return RiskCategory.INTEGRITY_RISK
    
    if risk.risk_type in ["inconsistent_value", "Inconsistent Value", "value_conflict", "Value Conflict"]:
        # Same label, different hardcoded values - update omission
        return RiskCategory.INTEGRITY_RISK
    
    if risk.risk_type in ["logic_alert", "Logic Alert"]:
        # Semantic oddities from Logic Translator
        return RiskCategory.INTEGRITY_RISK
    
    # Tab 3: Structural Debt (Maintenance Issues)
    if risk.risk_type in ["hidden_hardcode", "Hidden Hardcode"]:
        # Check consistency - inconsistent hardcodes are integrity risks
        if all_risks:
            is_consistent = check_hardcode_consistency(risk, all_risks)
            if not is_consistent:
                # Change risk type to "Inconsistent Value" for clarity
                # This distinguishes update omissions from maintenance issues
                risk.risk_type = "Inconsistent Value"
                return RiskCategory.INTEGRITY_RISK
        
        # Consistent hardcodes are structural debt
        return RiskCategory.STRUCTURAL_DEBT
    
    if risk.risk_type in ["merged_cell", "Merged Cell"]:
        return RiskCategory.STRUCTURAL_DEBT
    
    # Default to structural debt for unknown types
    return RiskCategory.STRUCTURAL_DEBT


def check_hardcode_consistency(risk: RiskAlert, all_risks: List[RiskAlert]) -> bool:
    """
    Check if a hardcode risk has consistent values across similar contexts.
    
    Returns True if consistent (Structural Debt), False if inconsistent (Integrity Risk).
    
    Logic:
    - Find all hardcodes with the same row label
    - If they all have the same value → Consistent (Structural Debt)
    - If they have different values → Inconsistent (Integrity Risk - update omission)
    
    Args:
        risk: The hardcode risk to check
        all_risks: All risks in the model
        
    Returns:
        True if consistent, False if inconsistent
    """
    if risk.risk_type not in ["hidden_hardcode", "Hidden Hardcode"]:
        return True
    
    # Get the hardcoded value from this risk
    current_value = risk.details.get("hardcoded_value")
    if current_value is None:
        return True  # Can't determine, assume consistent
    
    # Find all hardcodes with same row label
    same_label_risks = [
        r for r in all_risks 
        if r.risk_type in ["hidden_hardcode", "Hidden Hardcode"]
        and r.row_label == risk.row_label
        and r.row_label is not None  # Must have a label to compare
    ]
    
    if len(same_label_risks) <= 1:
        return True  # Only one instance, assume consistent
    
    # Check if all values are the same
    values = []
    for r in same_label_risks:
        val = r.details.get("hardcoded_value")
        if val is not None:
            # Normalize for comparison (handle floats)
            if isinstance(val, (int, float)):
                values.append(float(val))
            else:
                values.append(str(val))
    
    if not values:
        return True  # No values to compare
    
    # All values should be the same
    unique_values = set(values)
    return len(unique_values) == 1


class RiskTriageEngine:
    """
    Engine for classifying risks into 3-tier triage system.
    
    Organizes risks by business impact:
    - Tab 1: Fatal Errors (model is broken)
    - Tab 2: Integrity Risks (hidden bugs - HIGHEST PRIORITY)
    - Tab 3: Structural Debt (maintenance issues)
    """
    
    def __init__(self, risks: List[RiskAlert]):
        """
        Initialize triage engine with list of risks.
        
        Args:
            risks: List of all detected risks
        """
        self.risks = risks
        self.fatal_errors: List[RiskAlert] = []
        self.integrity_risks: List[RiskAlert] = []
        self.structural_debt: List[RiskAlert] = []
    
    def classify_all(self):
        """
        Classify all risks into three categories.
        
        Updates the category field on each risk and populates the three lists.
        """
        for risk in self.risks:
            # Classify the risk
            category = classify_risk(risk, self.risks)
            
            # Update the risk's category field
            risk.category = category
            
            # Add to appropriate list
            if category == RiskCategory.FATAL_ERROR:
                self.fatal_errors.append(risk)
            elif category == RiskCategory.INTEGRITY_RISK:
                self.integrity_risks.append(risk)
            else:
                self.structural_debt.append(risk)
    
    def get_tab_counts(self) -> Dict[str, int]:
        """
        Get risk counts for tab labels.
        
        Returns:
            Dictionary with counts: {"fatal": 3, "integrity": 7, "structural": 12}
        """
        return {
            "fatal": len(self.fatal_errors),
            "integrity": len(self.integrity_risks),
            "structural": len(self.structural_debt)
        }
    
    def get_total_count(self) -> int:
        """Get total number of risks"""
        return len(self.risks)
    
    def get_risks_by_category(self, category: RiskCategory) -> List[RiskAlert]:
        """
        Get all risks for a specific category.
        
        Args:
            category: The category to filter by
            
        Returns:
            List of risks in that category
        """
        if category == RiskCategory.FATAL_ERROR:
            return self.fatal_errors
        elif category == RiskCategory.INTEGRITY_RISK:
            return self.integrity_risks
        else:
            return self.structural_debt
