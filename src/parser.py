"""
Excel parser with Virtual Fill support

This module provides the ExcelParser class that converts Excel files into
ModelAnalysis objects with full formula extraction and dependency tracking.
"""

from io import BytesIO
from typing import Dict, List, Optional, Any
import signal
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import networkx as nx
import streamlit as st

from src.models import CellInfo, ModelAnalysis, RiskAlert


class TimeoutException(Exception):
    """Exception raised when parsing times out"""
    pass


def timeout_handler(signum, frame):
    """Handler for timeout signal"""
    raise TimeoutException("Parsing timeout exceeded")


class ExcelParser:
    """
    Parser for Excel files that extracts formulas, dependencies, and metadata.
    
    Key features:
    - Virtual Fill for merged cells
    - Formula dependency extraction
    - Dynamic reference detection (INDIRECT, OFFSET, ADDRESS)
    - Dependency graph construction
    """
    
    def __init__(self):
        """Initialize the parser"""
        pass
    
    def parse(self, file_obj: BytesIO, filename: str = "unknown.xlsx") -> ModelAnalysis:
        """
        Main entry point - parse an Excel file into a ModelAnalysis object.
        
        Args:
            file_obj: BytesIO object containing the Excel file
            filename: Name of the file for reference
            
        Returns:
            ModelAnalysis object containing all parsed data
            
        Raises:
            ValueError: If file is password-protected or corrupt
            TimeoutException: If parsing exceeds 60 seconds
            MemoryError: If file is too large
            Exception: For other parsing errors
        """
        try:
            # Set up timeout (60 seconds)
            # Note: signal.alarm only works on Unix systems
            # For Windows compatibility, we'll skip the timeout in this MVP
            try:
                signal.signal(signal.SIGALRM, timeout_handler)
                signal.alarm(60)  # 60 second timeout
            except (AttributeError, ValueError):
                # Windows doesn't support SIGALRM, skip timeout
                pass
            
            # Load workbook with data_only=False to get formulas
            workbook = openpyxl.load_workbook(file_obj, data_only=False)
            
            # Extract basic info
            sheets = workbook.sheetnames
            all_cells: Dict[str, CellInfo] = {}
            all_merged_ranges: Dict[str, List[str]] = {}
            
            # Process each sheet
            for sheet_name in sheets:
                worksheet = workbook[sheet_name]
                
                # Identify merged ranges
                merged_ranges = self._identify_merged_ranges(worksheet)
                if merged_ranges:
                    all_merged_ranges[sheet_name] = merged_ranges
                
                # Parse cells with Virtual Fill
                sheet_cells = self._parse_sheet(worksheet, sheet_name, merged_ranges)
                all_cells.update(sheet_cells)
            
            # Build dependency graph
            dependency_graph = self._build_dependency_graph(all_cells)
            
            # Create ModelAnalysis (risks and health_score will be added by analyzer)
            model = ModelAnalysis(
                filename=filename,
                sheets=sheets,
                cells=all_cells,
                risks=[],
                health_score=100,  # Default, will be calculated by analyzer
                dependency_graph=dependency_graph,
                merged_ranges=all_merged_ranges
            )
            
            # Cancel timeout
            try:
                signal.alarm(0)
            except (AttributeError, ValueError):
                pass
            
            return model
            
        except TimeoutException:
            # Cancel timeout
            try:
                signal.alarm(0)
            except (AttributeError, ValueError):
                pass
            raise ValueError("File too complex for MVP analysis. Parsing exceeded 60 seconds.")
        
        except MemoryError:
            # Cancel timeout
            try:
                signal.alarm(0)
            except (AttributeError, ValueError):
                pass
            raise ValueError("File too complex for MVP analysis. File is too large.")
        
        except openpyxl.utils.exceptions.InvalidFileException as e:
            # Cancel timeout
            try:
                signal.alarm(0)
            except (AttributeError, ValueError):
                pass
            if "password" in str(e).lower() or "encrypted" in str(e).lower():
                raise ValueError("This file is password protected. Please upload an unencrypted version.")
            else:
                raise ValueError("Unable to parse file. The file may be corrupted.")
        
        except Exception as e:
            # Cancel timeout
            try:
                signal.alarm(0)
            except (AttributeError, ValueError):
                pass
            # Catch-all for other errors
            raise Exception(f"Error parsing Excel file: {str(e)}")
    
    def _identify_merged_ranges(self, worksheet: Worksheet) -> List[str]:
        """
        Extract all merged cell ranges from a worksheet.
        
        Args:
            worksheet: openpyxl Worksheet object
            
        Returns:
            List of merged range strings (e.g., ["A1:B3", "D5:E6"])
        """
        merged_ranges = []
        for merged_range in worksheet.merged_cells.ranges:
            merged_ranges.append(str(merged_range))
        return merged_ranges
    
    def _parse_sheet(self, worksheet: Worksheet, sheet_name: str, 
                     merged_ranges: List[str]) -> Dict[str, CellInfo]:
        """
        Parse all cells in a worksheet, applying Virtual Fill for merged cells.
        
        Args:
            worksheet: openpyxl Worksheet object
            sheet_name: Name of the sheet
            merged_ranges: List of merged range strings
            
        Returns:
            Dictionary of CellInfo objects keyed by "Sheet!Address"
        """
        cells: Dict[str, CellInfo] = {}
        
        # Create a map of merged ranges for quick lookup
        merged_map: Dict[str, str] = {}  # Maps cell address to range string
        for range_str in merged_ranges:
            # Parse range (e.g., "A1:B3")
            cell_range = worksheet[range_str]
            
            # Handle single cell case (openpyxl returns Cell, not tuple)
            if hasattr(cell_range, 'coordinate'):
                # Single cell
                merged_map[cell_range.coordinate] = range_str
            else:
                # Multiple cells - iterate through rows
                for row in cell_range:
                    # Handle single row case
                    if hasattr(row, 'coordinate'):
                        merged_map[row.coordinate] = range_str
                    else:
                        for cell in row:
                            if cell.coordinate:
                                merged_map[cell.coordinate] = range_str
        
        # Limit parsing to reasonable dimensions to prevent hangs
        # Use the actual used range, but cap at 10,000 rows
        # For columns, we'll process all columns but only non-empty cells
        max_row = min(worksheet.max_row, 10000) if worksheet.max_row else 1000
        
        # Iterate through cells in the used range only
        # Use iter_rows without max_col to process all columns
        cell_count = 0
        for row in worksheet.iter_rows(max_row=max_row):
            for cell in row:
                if cell.coordinate is None:
                    continue
                
                # Skip completely empty cells (no value, no formula, not merged)
                if cell.value is None and cell.coordinate not in merged_map:
                    continue
                
                cell_count += 1
                # Safety limit: stop if we've processed too many cells
                # Increased limit to handle larger models
                if cell_count > 100000:
                    break
                
                address = cell.coordinate
                value = cell.value
                formula = None
                
                # Extract formula if present
                # Check both cell.value (for simple formulas) and cell.formula (for error cells)
                if hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                elif hasattr(cell, 'formula') and cell.formula:
                    # For cells with errors (#REF!, #DIV/0!, etc.), the formula is in cell.formula
                    formula = cell.formula if cell.formula.startswith('=') else f"={cell.formula}"
                
                # Check if cell is merged
                is_merged = address in merged_map
                merged_range = merged_map.get(address) if is_merged else None
                
                # For merged cells, apply Virtual Fill
                if is_merged and merged_range:
                    # Get the top-left cell of the merged range
                    top_left_coord = merged_range.split(':')[0]
                    if address != top_left_coord:
                        # Copy value/formula from top-left cell
                        top_left_cell = worksheet[top_left_coord]
                        value = top_left_cell.value
                        if hasattr(top_left_cell, 'value') and isinstance(top_left_cell.value, str) and top_left_cell.value.startswith('='):
                            formula = top_left_cell.value
                        elif hasattr(top_left_cell, 'formula') and top_left_cell.formula:
                            # For cells with errors, the formula is in cell.formula
                            formula = top_left_cell.formula if top_left_cell.formula.startswith('=') else f"={top_left_cell.formula}"
                
                # Extract dependencies if formula exists
                dependencies = []
                is_dynamic = False
                if formula:
                    dependencies = self._extract_dependencies(formula, sheet_name)
                    is_dynamic = self._is_dynamic_formula(formula)
                
                # Create CellInfo
                cell_info = CellInfo(
                    sheet=sheet_name,
                    address=address,
                    value=value,
                    formula=formula,
                    dependencies=dependencies,
                    is_dynamic=is_dynamic,
                    is_merged=is_merged,
                    merged_range=merged_range
                )
                
                # Store with full address as key
                key = f"{sheet_name}!{address}"
                cells[key] = cell_info
        
        return cells
    
    def _extract_dependencies(self, formula: str, current_sheet: str) -> List[str]:
        """
        Extract cell references from a formula using openpyxl tokenizer.
        
        Expands range references (A1:B10) into individual cells for accurate dependency tracking.
        
        Args:
            formula: Formula string (e.g., "=A1+B2" or "=SUM(A1:A10)")
            current_sheet: Name of the current sheet for relative references
            
        Returns:
            List of dependencies in "Sheet!Address" format
        """
        from openpyxl.formula.tokenizer import Tokenizer, Token
        from openpyxl.utils import range_boundaries, get_column_letter
        
        dependencies = []
        
        try:
            # Keep the '=' sign - the tokenizer needs it!
            formula_str = formula if formula.startswith('=') else f"={formula}"
            
            # Tokenize the formula
            tokenizer = Tokenizer(formula_str)
            tokens = tokenizer.items
            
            for token in tokens:
                # Look for cell references and ranges
                if token.type == Token.OPERAND:
                    subtype = token.subtype
                    value = token.value
                    
                    # Check if it's a cell reference or range
                    if subtype == Token.RANGE:
                        # Extract sheet name if present
                        sheet_name = current_sheet
                        cell_ref = value
                        
                        if '!' in value:
                            # Cross-sheet reference (e.g., "Sheet1!A1" or "Sheet1!A1:B10")
                            parts = value.split('!')
                            sheet_name = parts[0].strip("'")  # Remove quotes if present
                            cell_ref = parts[1]
                        
                        # Check if it's a range (contains ':')
                        if ':' in cell_ref:
                            # Expand range into individual cells
                            try:
                                min_col, min_row, max_col, max_row = range_boundaries(cell_ref)
                                
                                # Limit expansion to prevent memory issues (max 1000 cells)
                                total_cells = (max_col - min_col + 1) * (max_row - min_row + 1)
                                if total_cells > 1000:
                                    # For very large ranges, just add the range itself
                                    dependencies.append(f"{sheet_name}!{cell_ref}")
                                else:
                                    # Expand range into individual cells
                                    for row in range(min_row, max_row + 1):
                                        for col in range(min_col, max_col + 1):
                                            col_letter = get_column_letter(col)
                                            dependencies.append(f"{sheet_name}!{col_letter}{row}")
                            except:
                                # If expansion fails, add the range as-is
                                dependencies.append(f"{sheet_name}!{cell_ref}")
                        else:
                            # Single cell reference
                            dependencies.append(f"{sheet_name}!{cell_ref}")
        
        except Exception as e:
            # If tokenization fails, log the error for debugging
            print(f"[WARNING] Failed to extract dependencies from formula: {formula[:50]}... Error: {str(e)}")
        
        return dependencies
    
    def _is_dynamic_formula(self, formula: str) -> bool:
        """
        Check if formula contains dynamic reference functions.
        
        Args:
            formula: Formula string
            
        Returns:
            True if formula contains INDIRECT, OFFSET, or ADDRESS
        """
        # Convert to uppercase for case-insensitive matching
        formula_upper = formula.upper()
        
        # Check for dynamic functions
        dynamic_functions = ['INDIRECT(', 'OFFSET(', 'ADDRESS(']
        
        for func in dynamic_functions:
            if func in formula_upper:
                return True
        
        return False
    
    def _build_dependency_graph(self, cells: Dict[str, CellInfo]) -> nx.DiGraph:
        """
        Construct a NetworkX directed graph from cell dependencies.
        
        Args:
            cells: Dictionary of CellInfo objects
            
        Returns:
            NetworkX DiGraph with cells as nodes and dependencies as edges
        """
        graph = nx.DiGraph()
        
        # Add all cells as nodes
        for cell_key, cell_info in cells.items():
            graph.add_node(cell_key)
        
        # Add edges for dependencies
        for cell_key, cell_info in cells.items():
            # Skip dynamic formulas to prevent infinite recursion
            if cell_info.is_dynamic:
                continue
            
            # Add edge from each dependency to this cell
            for dep in cell_info.dependencies:
                # Only add edge if the dependency exists in our cells
                if dep in cells:
                    # Edge direction: dependency -> cell (dependency points to dependent)
                    graph.add_edge(dep, cell_key)
        
        return graph



@st.cache_data(show_spinner=False)
def parse_excel_file(file_bytes: bytes, filename: str) -> ModelAnalysis:
    """
    Cached wrapper for Excel parsing.
    
    This function is cached by Streamlit using the file content as the cache key.
    Re-uploading the same file will retrieve results from cache without re-parsing.
    
    Args:
        file_bytes: Raw bytes of the Excel file
        filename: Name of the file
        
    Returns:
        ModelAnalysis object
    """
    parser = ExcelParser()
    file_obj = BytesIO(file_bytes)
    return parser.parse(file_obj, filename)
