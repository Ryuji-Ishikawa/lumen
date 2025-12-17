"""
Smart Context Recovery - AI-powered context extraction

This module provides AI-powered context recovery for complex Excel layouts
that defeat rule-based parsing (2-column layouts, merged cells, etc.).

Strategy: Hybrid approach
- Rule-based parsing: 80% accuracy (fast, free)
- AI recovery: 20% fallback (accurate, costs API calls)
"""

from typing import Dict, Optional, Tuple
import re
from src.models import CellInfo


class SmartContextRecovery:
    """
    AI-powered context recovery for complex layouts.
    
    Uses LLM to understand 2-column layouts, merged cells,
    and other patterns that defeat rule-based parsing.
    """
    
    def __init__(self, ai_provider: str = "OpenAI", api_key: Optional[str] = None):
        """
        Initialize Smart Context Recovery.
        
        Args:
            ai_provider: AI provider ("OpenAI", "Google", etc.)
            api_key: API key for the provider
        """
        self.ai_provider = ai_provider
        self.api_key = api_key
        self.cache = {}
        self.enabled = api_key is not None
    
    def recover_context(self, sheet: str, cell_address: str, 
                       cells: Dict[str, CellInfo]) -> Optional[str]:
        """
        Use AI to recover context for a cell with surrounding context window.
        
        Args:
            sheet: Sheet name
            cell_address: Cell address (e.g., "E92")
            cells: Dictionary of all cells
            
        Returns:
            Recovered context label or None
        """
        if not self.enabled:
            return None
        
        # Extract context window (surrounding cells)
        context_window = self._extract_context_window(sheet, cell_address, cells)
        
        if not any(context_window.values()):
            return None
        
        # Check cache
        cache_key = self._make_cache_key_from_context(context_window)
        if cache_key in self.cache:
            return self.cache[cache_key]
        
        # Query LLM with context window
        label = self._query_llm_with_context(context_window, cell_address)
        
        # Cache result
        if label:
            self.cache[cache_key] = label
        
        return label
    
    def _extract_grid(self, sheet: str, cell_address: str, 
                     cells: Dict[str, CellInfo]) -> Dict[str, str]:
        """
        Extract 5x5 grid around target cell.
        
        Args:
            sheet: Sheet name
            cell_address: Cell address (e.g., "E92")
            cells: Dictionary of all cells
            
        Returns:
            Dictionary mapping cell addresses to masked values
        """
        from openpyxl.utils import column_index_from_string, get_column_letter
        
        # Parse cell address
        match = re.match(r'^([A-Z]+)(\d+)$', cell_address)
        if not match:
            return {}
        
        col_letter = match.group(1)
        row_num = int(match.group(2))
        col_num = column_index_from_string(col_letter)
        
        # Extract 5x5 grid (2 cells in each direction)
        grid = {}
        for row_offset in range(-2, 3):  # -2, -1, 0, 1, 2
            for col_offset in range(-2, 3):
                check_row = row_num + row_offset
                check_col = col_num + col_offset
                
                # Skip invalid cells
                if check_row < 1 or check_col < 1:
                    continue
                
                check_col_letter = get_column_letter(check_col)
                check_address = f"{check_col_letter}{check_row}"
                key = f"{sheet}!{check_address}"
                
                cell = cells.get(key)
                if cell and cell.value:
                    # Mask the value (security)
                    masked_value = self._mask_value(cell.value, cell.formula)
                    grid[check_address] = masked_value
                else:
                    grid[check_address] = "[EMPTY]"
        
        return grid
    
    def _mask_value(self, value, formula: Optional[str]) -> str:
        """
        Mask cell value for security.
        
        CRITICAL: Never send raw financial values to LLM.
        
        Args:
            value: Cell value
            formula: Cell formula (if any)
            
        Returns:
            Masked value
        """
        # If it's a formula, mask it
        if formula:
            return "[FORMULA]"
        
        # If it's a number, mask it
        if isinstance(value, (int, float)):
            return "[NUM]"
        
        # If it's a string that looks like a number, mask it
        if isinstance(value, str):
            try:
                float(value.replace(',', ''))
                return "[NUM]"
            except ValueError:
                pass
        
        # Keep text labels intact
        return str(value)
    
    def _make_cache_key(self, grid: Dict[str, str]) -> str:
        """
        Create cache key from grid.
        
        Args:
            grid: Dictionary mapping cell addresses to masked values
            
        Returns:
            Cache key string
        """
        # Sort by cell address for consistent key
        sorted_items = sorted(grid.items())
        return "|".join([f"{addr}:{val}" for addr, val in sorted_items])
    
    def _make_cache_key_from_context(self, context: Dict[str, list]) -> str:
        """
        Create cache key from context window.
        
        Args:
            context: Dictionary with "left", "above", "right", "below" lists
            
        Returns:
            Cache key string
        """
        parts = []
        for direction in ["left", "above", "right", "below"]:
            values = context.get(direction, [])
            parts.append(f"{direction}:{','.join(values)}")
        return "|".join(parts)
    
    def _extract_context_window(self, sheet: str, cell_address: str, 
                               cells: Dict[str, CellInfo]) -> Dict[str, list]:
        """
        Extract surrounding cells to provide context for AI recovery.
        
        Returns a context window with 3-5 cells in each direction.
        This gives the AI "eyes" to understand the layout.
        
        Args:
            sheet: Sheet name
            cell_address: Cell address (e.g., "E92")
            cells: Dictionary of all cells
            
        Returns:
            Dict with "left", "above", "right", "below" lists of cell values
        """
        from openpyxl.utils import column_index_from_string, get_column_letter
        
        # Parse cell address (e.g., "E92" -> col=E, row=92)
        match = re.match(r'^([A-Z]+)(\d+)$', cell_address)
        if not match:
            return {"left": [], "above": [], "right": [], "below": []}
        
        col_letter = match.group(1)
        row_num = int(match.group(2))
        col_num = column_index_from_string(col_letter)
        
        context = {
            "left": [],
            "above": [],
            "right": [],
            "below": []
        }
        
        # Get cells to the left (same row) - scan ALL the way to column A for labels
        # CRITICAL FIX: Scan to column A (not just 10 cells) to find item names
        # Limit to 30 text labels to avoid overwhelming the AI
        text_labels_found = 0
        max_labels = 30
        
        for i in range(1, col_num):  # Scan from current column to A
            if text_labels_found >= max_labels:
                break  # Stop after finding enough labels
            
            left_col_num = col_num - i
            if left_col_num < 1:
                break
            
            left_col = get_column_letter(left_col_num)
            left_cell = f"{sheet}!{left_col}{row_num}"
            
            if left_cell not in cells:
                continue  # Cell doesn't exist, keep scanning
            
            cell = cells[left_cell]
            value = cell.value
            
            if not value:
                continue  # Empty cell, keep scanning
            
            # CRITICAL FIX: Only add TEXT labels (skip numbers and formulas)
            # Skip formula strings like "=BJ24" - they're not useful for AI
            if isinstance(value, str):
                # Skip formula strings (displayed as text)
                if value.startswith('='):
                    continue  # Skip formula strings
                # Skip pure numbers as strings
                cleaned = value.replace('.', '').replace('-', '').replace(',', '').replace(' ', '').replace('+', '')
                if cleaned.isdigit():
                    continue  # Skip numeric strings
                # Skip whitespace-only strings
                if not value.strip():
                    continue
                # Accept text labels
                context["left"].append(str(value))
                text_labels_found += 1
            elif isinstance(value, (int, float)):
                # Skip numeric values - we only want text labels for context
                continue
            # For other types (datetime, etc.), skip them too
        
        # Get 5 cells above (same column)
        for i in range(1, 6):
            if row_num - i < 1:
                break
            above_cell = f"{sheet}!{col_letter}{row_num - i}"
            if above_cell in cells and cells[above_cell].value:
                context["above"].append(str(cells[above_cell].value))
        
        # Get 3 cells to the right (same row)
        for i in range(1, 4):
            right_col = get_column_letter(col_num + i)
            right_cell = f"{sheet}!{right_col}{row_num}"
            if right_cell in cells and cells[right_cell].value:
                context["right"].append(str(cells[right_cell].value))
        
        # Get 3 cells below (same column)
        for i in range(1, 4):
            below_cell = f"{sheet}!{col_letter}{row_num + i}"
            if below_cell in cells and cells[below_cell].value:
                context["below"].append(str(cells[below_cell].value))
        
        # Log context window for debugging
        print(f"[AI] Context window for {sheet}!{cell_address}:")
        print(f"  Left: {context['left']}")
        print(f"  Above: {context['above']}")
        print(f"  Right: {context['right']}")
        print(f"  Below: {context['below']}")
        
        return context
    
    def _query_llm_with_context(self, context_window: Dict[str, list], 
                                target_cell: str) -> Optional[str]:
        """
        Query LLM for semantic label using context window.
        
        Args:
            context_window: Dictionary with "left", "above", "right", "below" lists
            target_cell: Target cell address
            
        Returns:
            Semantic label or None
        """
        if not self.api_key:
            return None
        
        # Build prompt with context window
        prompt = self._build_context_prompt(context_window, target_cell)
        
        # Query based on provider
        if self.ai_provider == "OpenAI":
            return self._query_openai(prompt)
        elif self.ai_provider == "Google":
            return self._query_google(prompt)
        else:
            return None
    
    def _query_llm(self, grid: Dict[str, str], target_cell: str) -> Optional[str]:
        """
        Query LLM for semantic label (legacy method).
        
        Args:
            grid: Dictionary mapping cell addresses to masked values
            target_cell: Target cell address (center of grid)
            
        Returns:
            Semantic label or None
        """
        if not self.api_key:
            return None
        
        # Build prompt
        prompt = self._build_prompt(grid, target_cell)
        
        # Query based on provider
        if self.ai_provider == "OpenAI":
            return self._query_openai(prompt)
        elif self.ai_provider == "Google":
            return self._query_google(prompt)
        else:
            return None
    
    def _build_context_prompt(self, context_window: Dict[str, list], 
                             target_cell: str) -> str:
        """
        Build prompt for LLM with context window.
        
        Args:
            context_window: Dictionary with "left", "above", "right", "below" lists
            target_cell: Target cell address
            
        Returns:
            Prompt string
        """
        # Format context window
        left_str = ", ".join(context_window.get("left", [])) or "[empty]"
        above_str = ", ".join(context_window.get("above", [])) or "[empty]"
        right_str = ", ".join(context_window.get("right", [])) or "[empty]"
        below_str = ", ".join(context_window.get("below", [])) or "[empty]"
        
        prompt = f"""Role: Expert at understanding Japanese Excel financial models

Context:
- Cell: {target_cell}
- Context Window (surrounding cells):
  - Left (same row): {left_str}
  - Above (same column): {above_str}
  - Right (same row): {right_str}
  - Below (same column): {below_str}

Task: Look at the surrounding cells. If you find a clear text label, return it.

CRITICAL RULES:
1. ONLY return labels that actually exist in the surrounding cells shown above
2. Do NOT invent or guess labels based on the cell value
3. Do NOT make assumptions about what the data might represent
4. If you cannot find a clear text label in the surrounding cells, return "NONE"
5. If all surrounding cells are [empty] or contain only numbers, return "NONE"

Requirements:
- Return ONLY the label text that exists in the context, no explanation
- Be specific (e.g., "Cash Balance" not "Total")
- Use Japanese if the surrounding context is Japanese
- Maximum 50 characters
- If uncertain or no clear label exists, return "NONE"

Example:
Context: Above cells = ["Balance Sheet", "Current Assets", ""]
Good: "Current Assets" (exists in context)

Context: Left cells = ["Personnel Cost", "Marketing", "R&D"]
Good: "Personnel Cost" (exists in context)

Context: All cells = ["[empty]", "[empty]", "[empty]"]
Good: "NONE" (no labels found)

Answer:"""
        
        return prompt
    
    def _build_prompt(self, grid: Dict[str, str], target_cell: str) -> str:
        """
        Build prompt for LLM (legacy method).
        
        Args:
            grid: Dictionary mapping cell addresses to masked values
            target_cell: Target cell address
            
        Returns:
            Prompt string
        """
        # Format grid as table
        grid_str = "Excel Cell Grid:\n"
        for addr, val in sorted(grid.items()):
            marker = " ← TARGET" if addr == target_cell else ""
            grid_str += f"{addr}: {val}{marker}\n"
        
        prompt = f"""You are analyzing an Excel spreadsheet layout to find semantic labels.

{grid_str}

Question: Based on this layout, what is the semantic label (row header) for the TARGET cell?

Rules:
1. Look for text labels to the LEFT of the target cell
2. Ignore [NUM] and [FORMULA] cells
3. Return ONLY the label text, nothing else
4. If no clear label exists, return "NONE"

Answer:"""
        
        return prompt
    
    def _query_openai(self, prompt: str) -> Optional[str]:
        """
        Query OpenAI API.
        
        Args:
            prompt: Prompt string
            
        Returns:
            Response text or None
        """
        try:
            from openai import OpenAI
            
            client = OpenAI(api_key=self.api_key)
            
            response = client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": "You are an Excel layout analyzer."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=50,
                temperature=0
            )
            
            answer = response.choices[0].message.content.strip()
            
            # Clean up response
            if answer.upper() == "NONE":
                return None
            
            return answer
        
        except Exception as e:
            print(f"OpenAI API error: {str(e)}")
            return None
    
    def _query_google(self, prompt: str) -> Optional[str]:
        """
        Query Google AI API.
        
        Args:
            prompt: Prompt string
            
        Returns:
            Response text or None
        """
        try:
            import google.generativeai as genai
            
            genai.configure(api_key=self.api_key)
            model = genai.GenerativeModel('gemini-pro')
            
            response = model.generate_content(prompt)
            answer = response.text.strip()
            
            # Clean up response
            if answer.upper() == "NONE":
                return None
            
            return answer
        
        except Exception as e:
            print(f"Google AI API error: {str(e)}")
            return None
