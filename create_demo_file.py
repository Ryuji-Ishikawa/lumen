#!/usr/bin/env python3
"""
Create a demo Excel file with MANY intentional risks for recording
"""

import openpyxl
from openpyxl.styles import Font, PatternFill

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "予算管理"

# Header styling
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

# Create header
headers = ["項目", "予算額", "実績額", "差額", "進捗率", "担当者", "備考"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(1, col, header)
    cell.fill = header_fill
    cell.font = header_font

# Add data with MANY intentional risks
data = [
    ["人件費", 5000000, 4800000, "=B2-C2", "=C2/B2", "田中", "正常"],
    ["広告費", 3000000, 3200000, "=B3-C3", "=C3/B3", "佐藤", "予算超過"],
    # Inconsistent formula (risk!)
    ["設備費", 2000000, 1800000, 200000, "=C4/B4", "鈴木", ""],  # Wrong!
    ["交通費", 500000, 480000, "=B5-C5", "=C5/B5", "山田", ""],
    # Missing formula (risk!)
    ["通信費", 300000, 290000, 10000, 0.97, "高橋", ""],  # No formulas!
    ["消耗品費", 400000, 380000, "=B7-C7", "=C7/B7", "伊藤", ""],
    # Broken reference (risk!)
    ["水道光熱費", 600000, 620000, "=B8-C8", "=C8/Z8", "渡辺", ""],  # Invalid!
    # Inconsistent formula (risk!)
    ["研修費", 800000, 750000, 50000, "=C9/B9", "中村", ""],  # Wrong!
    ["福利厚生費", 1000000, 980000, "=B10-C10", "=C10/B10", "小林", ""],
    # Missing formula (risk!)
    ["雑費", 200000, 190000, 10000, 0.95, "加藤", ""],  # No formulas!
    ["会議費", 150000, 145000, "=B12-C12", "=C12/B12", "山本", ""],
    # Broken reference (risk!)
    ["印刷費", 250000, 240000, "=B13-C13", "=C13/AA13", "佐々木", ""],  # Invalid!
    # Inconsistent formula (risk!)
    ["旅費", 700000, 680000, 20000, "=C14/B14", "木村", ""],  # Wrong!
    ["保険料", 900000, 900000, "=B15-C15", "=C15/B15", "林", ""],
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row_idx, col_idx, value)

# Add summary with risks
ws.cell(17, 1, "合計")
ws.cell(17, 1).font = Font(bold=True)
ws.cell(17, 2, "=SUM(B2:B15)")
ws.cell(17, 3, "=SUM(C2:C15)")
# Inconsistent formula (risk!)
ws.cell(17, 4, "=B17-C17")
ws.cell(17, 5, "=C17/B17")

# Add sheet with external links (risk!)
ws2 = wb.create_sheet("参照データ")
ws2.cell(1, 1, "部門")
ws2.cell(1, 2, "予算")
ws2.cell(2, 1, "営業部")
ws2.cell(2, 2, "=予算管理!B2")
ws2.cell(3, 1, "外部データ1")
ws2.cell(3, 2, "='[OtherFile.xlsx]Sheet1'!A1")  # External link!
ws2.cell(4, 1, "外部データ2")
ws2.cell(4, 2, "='[Budget2024.xlsx]Data'!B5")  # External link!

# Adjust column widths
for col in range(1, 8):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

# Save file
filename = "Demo_Budget_With_Risks.xlsx"
wb.save(filename)

print(f"✅ Created: {filename}")
print()
print("📋 MANY intentional risks included:")
print("   - 5x Inconsistent formulas")
print("   - 3x Missing formulas")
print("   - 3x Broken references")
print("   - 2x External links")
print("   - Formula inconsistencies in summary")
print()
print("Total: 13+ risks for demo!")
