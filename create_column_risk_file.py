#!/usr/bin/env python3
"""
Create demo file with COLUMN-DIRECTION risks (what Lumen actually detects!)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "予算管理"

# Header
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
headers = ["項目", "予算額", "実績額", "差額", "進捗率", "担当者", "備考"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(1, col, header)
    cell.fill = header_fill
    cell.font = header_font

# Data with COLUMN-DIRECTION risks
# Most rows have =B-C for diff, =C/B for progress
# But some rows break the pattern!
data = [
    ["人件費", 5000000, 4800000, "=B2-C2", "=C2/B2", "田中", "正常"],
    ["広告費", 3000000, 3200000, "=B3-C3", "=C3/B3", "佐藤", ""],
    ["設備費", 2000000, 1800000, "=B4-C4", "=C4/B4", "鈴木", ""],
    ["交通費", 500000, 480000, "=B5-C5", "=C5/B5", "山田", ""],
    
    # RISK 1: Different formula in diff column (=C-B instead of =B-C)
    ["通信費", 300000, 290000, "=C6-B6", "=C6/B6", "高橋", "リスク1"],
    
    ["消耗品費", 400000, 380000, "=B7-C7", "=C7/B7", "伊藤", ""],
    
    # RISK 2: Broken reference in diff column
    ["水道光熱費", 600000, 620000, "=B8-Z8", "=C8/B8", "渡辺", "リスク2"],
    
    ["研修費", 800000, 750000, "=B9-C9", "=C9/B9", "中村", ""],
    ["福利厚生費", 1000000, 980000, "=B10-C10", "=C10/B10", "小林", ""],
    
    # RISK 3: No formula in diff column (hardcoded value)
    ["雑費", 200000, 190000, 10000, "=C11/B11", "加藤", "リスク3"],
    
    ["会議費", 150000, 145000, "=B12-C12", "=C12/B12", "山本", ""],
    
    # RISK 4: Different formula pattern in progress column (=B/C instead of =C/B)
    ["印刷費", 250000, 240000, "=B13-C13", "=B13/C13", "佐々木", "リスク4"],
    
    ["旅費", 700000, 680000, "=B14-C14", "=C14/B14", "木村", ""],
    
    # RISK 5: No formula in progress column
    ["保険料", 900000, 900000, "=B15-C15", 1.0, "林", "リスク5"],
    
    ["外注費", 1200000, 1150000, "=B16-C16", "=C16/B16", "松本", ""],
    
    # RISK 6: Broken reference in progress column
    ["リース料", 450000, 450000, "=B17-C17", "=C17/Z17", "井上", "リスク6"],
    
    ["修繕費", 350000, 340000, "=B18-C18", "=C18/B18", "木下", ""],
    
    # RISK 7: Different formula in diff column (=B2-C2 - wrong row reference!)
    ["広報費", 280000, 275000, "=B2-C2", "=C19/B19", "山口", "リスク7"],
    
    ["接待費", 320000, 310000, "=B20-C20", "=C20/B20", "斎藤", ""],
    
    # RISK 8: No formula in both columns
    ["交際費", 180000, 175000, 5000, 0.97, "清水", "リスク8"],
    
    ["寄付金", 100000, 100000, "=B22-C22", "=C22/B22", "森", ""],
    
    # RISK 9: Different formula pattern (=B+C instead of =B-C)
    ["諸経費", 220000, 215000, "=B23+C23", "=C23/B23", "池田", "リスク9"],
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row_idx, col_idx, value)

# Summary row (row 24, not 25!)
ws.cell(24, 1, "合計")
ws.cell(24, 1).font = Font(bold=True)
ws.cell(24, 2, "=SUM(B2:B23)")
ws.cell(24, 3, "=SUM(C2:C23)")
ws.cell(24, 4, "=B24-C24")
ws.cell(24, 5, "=C24/B24")

# Sheet 2 with external links
ws2 = wb.create_sheet("参照データ")
ws2.cell(1, 1, "部門")
ws2.cell(1, 2, "予算")
ws2.cell(2, 1, "営業部")
ws2.cell(2, 2, "=予算管理!B2")
# RISK 10-12: External links
ws2.cell(3, 1, "外部データ1")
ws2.cell(3, 2, "='[OtherFile.xlsx]Sheet1'!A1")
ws2.cell(4, 1, "外部データ2")
ws2.cell(4, 2, "='[Budget2024.xlsx]Data'!B5")
ws2.cell(5, 1, "外部データ3")
ws2.cell(5, 2, "='[Master.xlsx]Summary'!C10")

# Adjust widths
for col in range(1, 8):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

filename = "Demo_Budget_With_Risks.xlsx"
wb.save(filename)

print(f"✅ Created: {filename}")
print()
print("📋 COLUMN-DIRECTION RISKS (what Lumen detects!):")
print()
print("差額列 (D column) - should all be =B-C:")
print("   Row 6:  =C6-B6 (reversed)")
print("   Row 8:  =B8-Z8 (broken ref)")
print("   Row 11: 10000 (hardcoded)")
print("   Row 19: =B2-C2 (wrong row)")
print("   Row 21: 5000 (hardcoded)")
print("   Row 23: =B23+C23 (+ instead of -)")
print()
print("進捗率列 (E column) - should all be =C/B:")
print("   Row 13: =B13/C13 (reversed)")
print("   Row 15: 1.0 (hardcoded)")
print("   Row 17: =C17/Z17 (broken ref)")
print("   Row 21: 0.97 (hardcoded)")
print()
print("External links (Sheet 2):")
print("   3 external file references")
print()
print("Total: 12+ COLUMN-DIRECTION RISKS!")
