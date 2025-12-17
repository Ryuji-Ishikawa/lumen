#!/usr/bin/env python3
import openpyxl

wb = openpyxl.load_workbook('Sample_Business Plan.xlsx')
print(f"Sheets: {wb.sheetnames}")
print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"Sheet: {sheet_name}")
    print(f"  Rows: {ws.max_row}, Cols: {ws.max_column}")
    print(f"  Sample data (first 5 rows):")
    for row in range(1, min(6, ws.max_row + 1)):
        row_data = [ws.cell(row, col).value for col in range(1, min(8, ws.max_column + 1))]
        print(f"    Row {row}: {row_data}")
    print()
