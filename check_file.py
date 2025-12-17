#!/usr/bin/env python3
import openpyxl

wb = openpyxl.load_workbook('Demo_Budget_With_Risks.xlsx')
ws = wb.active

print(f"Rows: {ws.max_row}")
print(f"Sheets: {wb.sheetnames}")
print("\nSample data:")
for i in range(2, min(6, ws.max_row+1)):
    row_data = [ws.cell(i, j).value for j in range(1, 8)]
    print(f"Row {i}: {row_data}")
