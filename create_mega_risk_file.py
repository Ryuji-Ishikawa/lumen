#!/usr/bin/env python3
"""
Create MEGA RISK demo file with TONS of risks
"""

import openpyxl
from openpyxl.styles import Font, PatternFill

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "予算管理"

# Header
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
headers = ["項目", "予算額", "実績額", "差額", "進捗率", "担当者", "備考"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(1, col, header)
    cell.fill = header_fill
    cell.font = header_font

# MEGA DATA with TONS of risks
data = [
    # Normal rows
    ["人件費", 5000000, 4800000, "=B2-C2", "=C2/B2", "田中", "正常"],
    ["広告費", 3000000, 3200000, "=B3-C3", "=C3/B3", "佐藤", "予算超過"],
    
    # RISK 1: Inconsistent formula
    ["設備費", 2000000, 1800000, 200000, "=C4/B4", "鈴木", "リスク1"],
    
    ["交通費", 500000, 480000, "=B5-C5", "=C5/B5", "山田", ""],
    
    # RISK 2: Missing formulas
    ["通信費", 300000, 290000, 10000, 0.97, "高橋", "リスク2"],
    
    ["消耗品費", 400000, 380000, "=B7-C7", "=C7/B7", "伊藤", ""],
    
    # RISK 3: Broken reference
    ["水道光熱費", 600000, 620000, "=B8-C8", "=C8/Z8", "渡辺", "リスク3"],
    
    # RISK 4: Inconsistent formula
    ["研修費", 800000, 750000, 50000, "=C9/B9", "中村", "リスク4"],
    
    ["福利厚生費", 1000000, 980000, "=B10-C10", "=C10/B10", "小林", ""],
    
    # RISK 5: Missing formulas
    ["雑費", 200000, 190000, 10000, 0.95, "加藤", "リスク5"],
    
    ["会議費", 150000, 145000, "=B12-C12", "=C12/B12", "山本", ""],
    
    # RISK 6: Broken reference
    ["印刷費", 250000, 240000, "=B13-C13", "=C13/AA13", "佐々木", "リスク6"],
    
    # RISK 7: Inconsistent formula
    ["旅費", 700000, 680000, 20000, "=C14/B14", "木村", "リスク7"],
    
    ["保険料", 900000, 900000, "=B15-C15", "=C15/B15", "林", ""],
    
    # RISK 8: Inconsistent formula
    ["外注費", 1200000, 1150000, 50000, "=C16/B16", "松本", "リスク8"],
    
    ["リース料", 450000, 450000, "=B17-C17", "=C17/B17", "井上", ""],
    
    # RISK 9: Missing formulas
    ["修繕費", 350000, 340000, 10000, 0.97, "木下", "リスク9"],
    
    ["広報費", 280000, 275000, "=B19-C19", "=C19/B19", "山口", ""],
    
    # RISK 10: Broken reference
    ["接待費", 320000, 310000, "=B20-C20", "=C20/ZZ20", "斎藤", "リスク10"],
    
    # RISK 11: Inconsistent formula
    ["交際費", 180000, 175000, 5000, "=C21/B21", "清水", "リスク11"],
    
    ["寄付金", 100000, 100000, "=B22-C22", "=C22/B22", "森", ""],
    
    # RISK 12: Missing formulas
    ["諸経費", 220000, 215000, 5000, 0.98, "池田", "リスク12"],
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row_idx, col_idx, value)

# Summary with RISK 13: Inconsistent formula
ws.cell(25, 1, "合計")
ws.cell(25, 1).font = Font(bold=True)
ws.cell(25, 2, "=SUM(B2:B23)")
ws.cell(25, 3, "=SUM(C2:C23)")
ws.cell(25, 4, "=B25-C25")  # RISK 13!
ws.cell(25, 5, "=C25/B25")

# Sheet 2 with external links
ws2 = wb.create_sheet("参照データ")
ws2.cell(1, 1, "部門")
ws2.cell(1, 2, "予算")
ws2.cell(2, 1, "営業部")
ws2.cell(2, 2, "=予算管理!B2")
# RISK 14: External link
ws2.cell(3, 1, "外部データ1")
ws2.cell(3, 2, "='[OtherFile.xlsx]Sheet1'!A1")
# RISK 15: External link
ws2.cell(4, 1, "外部データ2")
ws2.cell(4, 2, "='[Budget2024.xlsx]Data'!B5")
# RISK 16: External link
ws2.cell(5, 1, "外部データ3")
ws2.cell(5, 2, "='[Master.xlsx]Summary'!C10")

# Adjust widths
for col in range(1, 8):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

filename = "Demo_Budget_With_Risks.xlsx"
wb.save(filename)

print(f"✅ Created: {filename}")
print()
print("📋 MEGA RISKS included:")
print("   - 6x Inconsistent formulas (rows 4, 9, 14, 16, 21, 25)")
print("   - 4x Missing formulas (rows 6, 11, 18, 23)")
print("   - 3x Broken references (rows 8, 13, 20)")
print("   - 3x External links (sheet 2)")
print()
print("Total: 16+ MEGA RISKS!")
