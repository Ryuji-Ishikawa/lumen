"""
Driver X-Ray Test Suite

Tests the ability to trace dependencies from hardcoded cells to ultimate drivers.
This is critical for understanding the impact of changes.
"""

import pytest
import openpyxl
from pathlib import Path
from io import BytesIO
import sys

sys.path.insert(0, str(Path(__file__).parent.parent))
from src.parser import ExcelParser
from src.analyzer import ModelAnalyzer


class TestDriverXRay:
    """Test suite for Driver X-Ray functionality"""
    
    def setup_method(self):
        """Setup for each test"""
        self.parser = ExcelParser()
        self.analyzer = ModelAnalyzer()
    
    def create_test_file_with_drivers(self):
        """Create test file with clear dependency chain"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PL"
        
        # Simple P&L structure
        ws['A1'] = 'Item'
        ws['B1'] = 'Value'
        
        # Hardcoded input
        ws['B2'] = 1000  # Hardcoded revenue base
        ws['A2'] = 'Revenue Base'
        
        # Intermediate calculations
        ws['B3'] = '=B2*1.1'  # Revenue (depends on B2)
        ws['A3'] = 'Revenue'
        
        ws['B4'] = '=B3*0.6'  # COGS (depends on B3)
        ws['A4'] = 'COGS'
        
        ws['B5'] = '=B3-B4'  # Gross Profit (depends on B3, B4)
        ws['A5'] = 'Gross Profit'
        
        ws['B6'] = 200  # Hardcoded operating expenses
        ws['A6'] = 'OpEx'
        
        ws['B7'] = '=B5-B6'  # EBITDA (depends on B5, B6) - DRIVER
        ws['A7'] = 'EBITDA'
        
        # Save to BytesIO
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)
        return file_obj
    
    def test_get_precedents(self):
        """Test that we can get cells a cell depends on"""
        file = self.create_test_file_with_drivers()
        model = self.parser.parse(file, 'test.xlsx')
        model = self.analyzer.analyze(model)
        
        # B5 (Gross Profit) should depend on B3 and B4
        precedents = model.get_precedents('PL!B5')
        
        print(f"\n✓ Precedents of PL!B5 (Gross Profit):")
        for prec in precedents:
            print(f"  - {prec}")
        
        assert 'PL!B3' in precedents, "Should depend on Revenue (B3)"
        assert 'PL!B4' in precedents, "Should depend on COGS (B4)"
    
    def test_get_dependents(self):
        """Test that we can get cells that depend on a cell"""
        file = self.create_test_file_with_drivers()
        model = self.parser.parse(file, 'test.xlsx')
        model = self.analyzer.analyze(model)
        
        # B3 (Revenue) should be used by B4 and B5
        dependents = model.get_dependents('PL!B3')
        
        print(f"\n✓ Dependents of PL!B3 (Revenue):")
        for dep in dependents:
            print(f"  - {dep}")
        
        assert 'PL!B4' in dependents, "COGS (B4) should depend on Revenue"
        assert 'PL!B5' in dependents, "Gross Profit (B5) should depend on Revenue"
    
    def test_trace_to_drivers(self):
        """Test tracing from hardcoded cell to ultimate drivers"""
        file = self.create_test_file_with_drivers()
        model = self.parser.parse(file, 'test.xlsx')
        model = self.analyzer.analyze(model)
        
        # Trace from B2 (hardcoded revenue base) to drivers
        drivers = self.analyzer.trace_to_drivers(model, 'PL!B2')
        
        print(f"\n✓ Drivers affected by PL!B2 (Revenue Base):")
        for driver in drivers:
            cell = model.get_cell('PL', driver.split('!')[1])
            label = cell.value if cell else "Unknown"
            print(f"  - {driver}: {label}")
        
        # B7 (EBITDA) should be the ultimate driver
        assert 'PL!B7' in drivers, "EBITDA (B7) should be an ultimate driver"
        
        # B2 affects B7 through the chain: B2 → B3 → B4 → B5 → B7
        print(f"✓ Traced {len(drivers)} driver(s) from hardcoded cell")
    
    def test_trace_multiple_drivers(self):
        """Test tracing to multiple ultimate drivers"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Model"
        
        # Hardcoded input
        ws['A1'] = 100  # Hardcoded value
        
        # Multiple branches
        ws['A2'] = '=A1*2'  # Branch 1
        ws['A3'] = '=A1*3'  # Branch 2
        
        ws['A4'] = '=A2+10'  # Driver 1 (no dependents)
        ws['A5'] = '=A3+20'  # Driver 2 (no dependents)
        
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)
        
        model = self.parser.parse(file_obj, 'multi.xlsx')
        model = self.analyzer.analyze(model)
        
        # Trace from A1 to all drivers
        drivers = self.analyzer.trace_to_drivers(model, 'Model!A1')
        
        print(f"\n✓ Multiple drivers from Model!A1:")
        for driver in drivers:
            print(f"  - {driver}")
        
        # Should find both A4 and A5 as drivers
        assert 'Model!A4' in drivers, "A4 should be a driver"
        assert 'Model!A5' in drivers, "A5 should be a driver"
        assert len(drivers) == 2, "Should have exactly 2 drivers"
        
        print("✓ Correctly traced to multiple drivers")


if __name__ == '__main__':
    pytest.main([__file__, '-v', '-s'])
