"""
Driver X-Ray with Virtual Fill Test Suite

Tests that Driver X-Ray works correctly with merged cells (Virtual Fill).
This is critical per business owner's requirement: "The trace must work even 
if the driver is inside a merged cell."
"""

import pytest
import openpyxl
from pathlib import Path
from io import BytesIO
import sys

sys.path.insert(0, str(Path(__file__).parent.parent))
from src.parser import ExcelParser
from src.analyzer import ModelAnalyzer


class TestDriverXRayVirtualFill:
    """Test Driver X-Ray with merged cells"""
    
    def setup_method(self):
        """Setup for each test"""
        self.parser = ExcelParser()
        self.analyzer = ModelAnalyzer()
    
    def create_file_with_merged_driver(self):
        """Create file where the driver is in a merged cell"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dashboard"
        
        # Hardcoded input
        ws['A1'] = 1000
        
        # Intermediate calculation
        ws['A2'] = '=A1*1.5'
        
        # Driver in merged range B3:D3
        ws.merge_cells('B3:D3')
        ws['B3'] = '=A2+500'  # This is the driver (EBITDA)
        ws['A3'] = 'EBITDA'
        
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)
        return file_obj
    
    def test_trace_to_merged_driver(self):
        """Test tracing to a driver that's in a merged cell"""
        file = self.create_file_with_merged_driver()
        model = self.parser.parse(file, 'merged.xlsx')
        model = self.analyzer.analyze(model)
        
        print(f"\n✓ Merged ranges: {model.merged_ranges}")
        
        # Trace from A1 (hardcoded) to drivers
        drivers = self.analyzer.trace_to_drivers(model, 'Dashboard!A1')
        
        print(f"\n✓ Drivers affected by Dashboard!A1:")
        for driver in drivers:
            cell = model.cells.get(driver)
            if cell:
                print(f"  - {driver}: is_merged={cell.is_merged}, range={cell.merged_range}")
        
        # Should find B3 (and possibly virtual cells C3, D3)
        driver_addresses = [d.split('!')[1] for d in drivers]
        
        # At minimum, B3 should be a driver
        assert any('B3' in addr for addr in driver_addresses), "B3 should be a driver"
        
        print(f"✓ Successfully traced to driver in merged range")
        print(f"✓ Found {len(drivers)} driver cell(s)")
    
    def test_dependency_through_merged_cells(self):
        """Test that dependencies work correctly through merged cells"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Model"
        
        # Merged input range A1:C1
        ws.merge_cells('A1:C1')
        ws['A1'] = 100
        
        # Formula referencing merged range
        ws['A2'] = '=A1*2'  # References top-left
        ws['A3'] = '=B1*3'  # References middle (virtual cell)
        ws['A4'] = '=C1*4'  # References right (virtual cell)
        
        # Drivers
        ws['A5'] = '=A2+A3+A4'  # Ultimate driver
        
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)
        
        model = self.parser.parse(file_obj, 'virtual.xlsx')
        model = self.analyzer.analyze(model)
        
        # Check that virtual cells exist
        assert 'Model!A1' in model.cells, "A1 should exist"
        assert 'Model!B1' in model.cells, "B1 (virtual) should exist"
        assert 'Model!C1' in model.cells, "C1 (virtual) should exist"
        
        print(f"\n✓ Virtual cells created:")
        for addr in ['A1', 'B1', 'C1']:
            cell = model.cells.get(f'Model!{addr}')
            if cell:
                print(f"  - {addr}: is_merged={cell.is_merged}, value={cell.value}")
        
        # Trace from A1 to drivers
        drivers = self.analyzer.trace_to_drivers(model, 'Model!A1')
        
        print(f"\n✓ Drivers from Model!A1:")
        for driver in drivers:
            print(f"  - {driver}")
        
        # A5 should be the ultimate driver
        assert 'Model!A5' in drivers, "A5 should be the ultimate driver"
        
        print("✓ Dependencies work correctly through virtual cells")


if __name__ == '__main__':
    pytest.main([__file__, '-v', '-s'])
