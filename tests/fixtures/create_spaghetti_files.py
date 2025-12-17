"""
Create Spaghetti Excel Test Suite

This script generates test Excel files with various Japanese Excel patterns
that are known to break global tools. These files validate our Virtual Fill
algorithm and error handling robustness.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def create_heavy_merged_cells():
    """
    Test File 1: Heavy Merged Cells
    Pattern: Headers with 5+ merged columns (common in Japanese reports)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "財務諸表"  # Financial Statement
    
    # Merged header row 1: Company name across 12 columns
    ws.merge_cells('A1:L1')
    ws['A1'] = '株式会社サンプル 財務諸表'
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Merged header row 2: Period across 12 columns
    ws.merge_cells('A2:L2')
    ws['A2'] = '2024年度 第1四半期'
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Merged header row 3: Category headers
    ws.merge_cells('A3:A4')
    ws['A3'] = '勘定科目'  # Account Name
    ws.merge_cells('B3:D3')
    ws['B3'] = '第1四半期'  # Q1
    ws.merge_cells('E3:G3')
    ws['E3'] = '第2四半期'  # Q2
    ws.merge_cells('H3:J3')
    ws['H3'] = '第3四半期'  # Q3
    ws.merge_cells('K3:L3')
    ws['K3'] = '合計'  # Total
    
    # Sub-headers for Q1, Q2, Q3
    for col_start in ['B', 'E', 'H']:
        col_idx = ord(col_start) - ord('A')
        ws.cell(4, col_idx + 1, '予算')  # Budget
        ws.cell(4, col_idx + 2, '実績')  # Actual
        ws.cell(4, col_idx + 3, '差異')  # Variance
    
    ws.cell(4, 11, '予算')
    ws.cell(4, 12, '実績')
    
    # Data rows with formulas
    accounts = ['売上高', '売上原価', '販売費', '営業利益']
    for i, account in enumerate(accounts, start=5):
        ws.cell(i, 1, account)
        # Add some formulas with hardcoded values
        ws.cell(i, 2, f'=1000000*{i}')  # Hardcoded multiplier
        ws.cell(i, 3, f'=B{i}*1.1')  # Hardcoded 1.1
        ws.cell(i, 4, f'=C{i}-B{i}')
    
    wb.save('tests/fixtures/spaghetti_excel/heavy_merged_cells.xlsx')
    print("✓ Created: heavy_merged_cells.xlsx")


def create_complex_grid_layout():
    """
    Test File 2: Complex Grid Layout
    Pattern: Mixed merged rows and columns (Japanese planning sheets)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "計画シート"
    
    # Vertical merged cells for categories
    ws.merge_cells('A1:A5')
    ws['A1'] = '人件費'  # Personnel Cost
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A6:A10')
    ws['A6'] = '販売費'  # Sales Cost
    ws['A6'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Horizontal merged cells for sub-categories
    ws.merge_cells('B1:D1')
    ws['B1'] = '正社員'  # Full-time
    ws.merge_cells('B6:D6')
    ws['B6'] = '広告費'  # Advertising
    
    # Data with formulas referencing merged cells
    ws['B2'] = '基本給'  # Base salary
    ws['C2'] = 5000000
    ws['D2'] = '=C2*12'  # Hardcoded 12 months
    
    ws['B3'] = '賞与'  # Bonus
    ws['C3'] = '=C2*2.5'  # Hardcoded 2.5 months
    
    # Formula referencing merged range
    ws['E1'] = '=SUM(C1:D5)'  # References merged cells
    
    wb.save('tests/fixtures/spaghetti_excel/complex_grid_layout.xlsx')
    print("✓ Created: complex_grid_layout.xlsx")


def create_japanese_text_mixed():
    """
    Test File 3: Japanese + English Mixed
    Pattern: Mixed Japanese/English in formulas and cell references
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MixedText"
    
    # Headers
    ws['A1'] = 'Account Name'
    ws['B1'] = '予算 (Budget)'
    ws['C1'] = '実績 (Actual)'
    ws['D1'] = 'Variance'
    
    # Data with Japanese text
    ws['A2'] = '売上高 (Revenue)'
    ws['B2'] = 10000000
    ws['C2'] = 12000000
    ws['D2'] = '=C2-B2'
    
    ws['A3'] = 'Personnel Cost (人件費)'
    ws['B3'] = 5000000
    ws['C3'] = '=B3*1.05'  # Hardcoded 1.05
    ws['D3'] = '=C3-B3'
    
    # Formula with Japanese sheet reference (if we had multiple sheets)
    ws['A4'] = 'Total (合計)'
    ws['B4'] = '=SUM(B2:B3)'
    ws['C4'] = '=SUM(C2:C3)'
    ws['D4'] = '=C4-B4'
    
    wb.save('tests/fixtures/spaghetti_excel/japanese_text_mixed.xlsx')
    print("✓ Created: japanese_text_mixed.xlsx")


def create_circular_references():
    """
    Test File 4: Circular References
    Pattern: Intentional circular refs (common in iterative calculations)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Circular"
    
    ws['A1'] = 'Value A'
    ws['A2'] = 'Value B'
    ws['A3'] = 'Value C'
    
    # Create circular reference: A → B → C → A
    ws['B1'] = '=B3+100'  # Depends on C
    ws['B2'] = '=B1*0.5'  # Depends on A
    ws['B3'] = '=B2+50'   # Depends on B, creates cycle
    
    # Another circular reference
    ws['D1'] = '=D2+1'
    ws['D2'] = '=D1-1'
    
    wb.save('tests/fixtures/spaghetti_excel/circular_references.xlsx')
    print("✓ Created: circular_references.xlsx")


def create_cross_sheet_complex():
    """
    Test File 5: Cross-Sheet Complex
    Pattern: 10+ sheets with cross-references
    """
    wb = openpyxl.Workbook()
    
    # Create multiple sheets
    sheet_names = ['売上', '原価', '販売費', '管理費', '営業利益', 
                   '営業外収益', '営業外費用', '経常利益', '特別損益', '当期純利益']
    
    for i, name in enumerate(sheet_names):
        if i == 0:
            ws = wb.active
            ws.title = name
        else:
            ws = wb.create_sheet(name)
        
        ws['A1'] = f'{name}明細'
        ws['B1'] = '金額'
        ws['B2'] = (i + 1) * 1000000
        
        # Cross-sheet references
        if i > 0:
            prev_sheet = sheet_names[i-1]
            ws['B3'] = f"='{prev_sheet}'!B2*1.1"  # Reference previous sheet
    
    # Summary sheet with references to multiple sheets
    summary = wb.create_sheet('サマリー')
    summary['A1'] = '項目'
    summary['B1'] = '金額'
    
    for i, name in enumerate(sheet_names[:5], start=2):
        summary[f'A{i}'] = name
        summary[f'B{i}'] = f"='{name}'!B2"
    
    summary['A7'] = '合計'
    summary['B7'] = '=SUM(B2:B6)'
    
    wb.save('tests/fixtures/spaghetti_excel/cross_sheet_complex.xlsx')
    print("✓ Created: cross_sheet_complex.xlsx")


def create_edge_cases():
    """
    Test File 6: Edge Cases
    Pattern: Single cell merge, entire row merge, etc.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EdgeCases"
    
    # Edge case 1: Single cell "merge" (A1:A1)
    ws.merge_cells('A1:A1')
    ws['A1'] = 'Single Cell Merge'
    
    # Edge case 2: Entire row merge (very wide)
    ws.merge_cells('A3:Z3')
    ws['A3'] = '全行マージ (Entire Row Merged)'
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Edge case 3: Entire column merge (very tall)
    ws.merge_cells('B5:B20')
    ws['B5'] = '全列マージ'
    ws['B5'].alignment = Alignment(vertical='center')
    
    # Edge case 4: Large rectangular merge
    ws.merge_cells('D5:J15')
    ws['D5'] = '大きな矩形マージ'
    ws['D5'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Edge case 5: Adjacent merged cells
    ws.merge_cells('L5:M5')
    ws['L5'] = 'Merge 1'
    ws.merge_cells('N5:O5')
    ws['N5'] = 'Merge 2'
    
    # Formula referencing merged cells
    ws['A20'] = '=SUM(D5:J15)'  # References large merged area
    
    wb.save('tests/fixtures/spaghetti_excel/edge_cases.xlsx')
    print("✓ Created: edge_cases.xlsx")


def create_hanko_boxes():
    """
    Test File 7: Hanko Boxes (Japanese Approval Stamps)
    Pattern: Grid of merged cells for approval stamps (very common in Japanese Excel)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "承認欄"  # Approval Section
    
    # Main content area
    ws['A1'] = '申請書'  # Application Form
    ws['A2'] = '申請日'  # Application Date
    ws['B2'] = '2024-04-01'
    ws['A3'] = '申請者'  # Applicant
    ws['B3'] = '山田太郎'
    
    # Hanko box grid (top right corner - typical Japanese layout)
    # This is a 3x3 grid of merged cells for stamps
    hanko_start_row = 1
    hanko_start_col = 10  # Column J
    
    positions = ['部長', '課長', '担当']  # Manager, Section Chief, Staff
    
    for i, position in enumerate(positions):
        col = hanko_start_col + i * 2
        col_letter = get_column_letter(col)
        next_col_letter = get_column_letter(col + 1)
        
        # Position label (merged 2 columns)
        ws.merge_cells(f'{col_letter}{hanko_start_row}:{next_col_letter}{hanko_start_row}')
        ws[f'{col_letter}{hanko_start_row}'] = position
        ws[f'{col_letter}{hanko_start_row}'].alignment = Alignment(horizontal='center')
        ws[f'{col_letter}{hanko_start_row}'].fill = PatternFill(start_color='CCCCCC', fill_type='solid')
        
        # Stamp box (merged 2x2 for stamp area)
        ws.merge_cells(f'{col_letter}{hanko_start_row+1}:{next_col_letter}{hanko_start_row+2}')
        ws[f'{col_letter}{hanko_start_row+1}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Date row (merged 2 columns)
        ws.merge_cells(f'{col_letter}{hanko_start_row+3}:{next_col_letter}{hanko_start_row+3}')
        ws[f'{col_letter}{hanko_start_row+3}'] = '2024/04/01'
        ws[f'{col_letter}{hanko_start_row+3}'].alignment = Alignment(horizontal='center')
        ws[f'{col_letter}{hanko_start_row+3}'].font = Font(size=8)
    
    # Add some data with formulas below
    ws['A10'] = '項目'
    ws['B10'] = '金額'
    ws['A11'] = '交通費'
    ws['B11'] = 5000
    ws['A12'] = '宿泊費'
    ws['B12'] = 15000
    ws['A13'] = '合計'
    ws['B13'] = '=SUM(B11:B12)'
    
    wb.save('tests/fixtures/spaghetti_excel/hanko_boxes.xlsx')
    print("✓ Created: hanko_boxes.xlsx")


def create_test_documentation():
    """Create README documenting expected behavior for each test file"""
    doc = """# Spaghetti Excel Test Suite

This directory contains test Excel files with various Japanese Excel patterns that are known to break global tools.

## Test Files

### 1. heavy_merged_cells.xlsx
**Pattern**: Headers with 5+ merged columns
**Challenge**: Multiple levels of merged headers (company name, period, categories)
**Expected Behavior**: 
- Parser should identify all merged ranges
- Virtual Fill should propagate values to all coordinates
- Formulas with hardcoded values (1000000, 1.1) should be detected
- Dependency graph should include all virtual cells

### 2. complex_grid_layout.xlsx
**Pattern**: Mixed merged rows and columns
**Challenge**: Vertical merges for categories + horizontal merges for sub-categories
**Expected Behavior**:
- Parser should handle both vertical and horizontal merges
- Formulas referencing merged ranges should resolve correctly
- No cells should be lost or duplicated

### 3. japanese_text_mixed.xlsx
**Pattern**: Mixed Japanese/English text
**Challenge**: Japanese characters in cell values and formulas
**Expected Behavior**:
- Parser should handle UTF-8 Japanese text correctly
- Context labels should display Japanese characters
- No encoding errors

### 4. circular_references.xlsx
**Pattern**: Intentional circular references
**Challenge**: A → B → C → A cycle
**Expected Behavior**:
- Parser should not hang or crash
- Circular reference detector should identify all cycles
- Critical severity alerts should be created

### 5. cross_sheet_complex.xlsx
**Pattern**: 10+ sheets with cross-references
**Challenge**: Complex cross-sheet dependencies
**Expected Behavior**:
- Parser should handle all sheets
- Cross-sheet references should be extracted correctly
- Dependency graph should span multiple sheets
- Cross-sheet spaghetti detector should flag formulas with >2 external sheets

### 6. edge_cases.xlsx
**Pattern**: Extreme merge scenarios
**Challenge**: Single cell merge, entire row merge, entire column merge, large rectangular merge
**Expected Behavior**:
- Parser should handle all edge cases without crashing
- Virtual Fill should work for all merge types
- No performance degradation

### 7. hanko_boxes.xlsx
**Pattern**: Japanese approval stamp grid
**Challenge**: 3x3 grid of merged cells (very common in Japanese business Excel)
**Expected Behavior**:
- Parser should handle hanko box grid correctly
- Merged cells in top-right corner should not interfere with data area
- All formulas should be extracted correctly

## Success Criteria

For the parser to pass this test suite:
1. **No Crashes**: All files must parse without exceptions
2. **Specific Errors**: If parsing fails, error messages must be specific and actionable
3. **Partial Success**: If some cells fail, parser should continue and return partial results
4. **Virtual Fill**: All merged cells must have propagated values
5. **Dependency Graph**: All cells (including virtual) must be in the graph
6. **Performance**: Parsing should complete within 60 seconds per file

## Running Tests

```bash
pytest tests/test_parser_robustness.py -v
```

## Adding New Test Files

When adding new spaghetti patterns:
1. Create the Excel file in this directory
2. Document the pattern and expected behavior in this README
3. Add test case to test_parser_robustness.py
"""
    
    with open('tests/fixtures/spaghetti_excel/README.md', 'w', encoding='utf-8') as f:
        f.write(doc)
    
    print("✓ Created: README.md")


if __name__ == '__main__':
    print("Creating Spaghetti Excel Test Suite...")
    print("=" * 60)
    
    create_heavy_merged_cells()
    create_complex_grid_layout()
    create_japanese_text_mixed()
    create_circular_references()
    create_cross_sheet_complex()
    create_edge_cases()
    create_hanko_boxes()
    create_test_documentation()
    
    print("=" * 60)
    print("✅ Spaghetti Excel Test Suite created successfully!")
    print("\nNext step: Run 'python tests/fixtures/create_spaghetti_files.py'")
