"""
Context Type Filter Test

CRITICAL: Validates that context extraction ONLY picks TEXT labels,
not formulas or numbers.

UAT Failure: Context contained "=(D18*E18)" - This is WRONG.
Fix: Type filter to accept only text values.
"""

import pytest
import sys
from pathlib import Path
from io import BytesIO
import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent))
from src.parser import ExcelParser
from src.analyzer import ModelAnalyzer


class TestContextTypeFilter:
    """Test type filtering for context extraction"""
    
    def setup_method(self):
        """Setup for each test"""
        self.parser = ExcelParser()
        self.analyzer = ModelAnalyzer()
    
    def test_reject_formulas_as_context(self):
        """CRITICAL: Context should NOT be formulas"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test"
        
        # Layout: Formula in B5, Text label in A5, Target in C5
        ws['A5'] = '純資産'  # Text label (ACCEPT)
        ws['B5'] = '=D18*E18'  # Formula (REJECT)
        ws['C5'] = 1000  # Target cell
        
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)
        
        model = self.parser.parse(file_obj, 'test.xlsx')
        
        # Get context for C5
        row_label, col_label = self.analyzer._get_context_labels('Test', 'C5', model.cells)
        
        print(f"\n✓ Context for C5:")
        print(f"  Row Label: {row_label}")
        
        # CRITICAL: Should find A5 (純資産), NOT B5 (=D18*E18)
        assert row_label == '純資産', f"Should find text label, got: {row_label}"
        assert row_label != '=D18*E18', "Should NOT return formula as context"
        
        print("✓ PASS: Formula rejected, text label found")
    
    def test_reject_numbers_as_context(self):
        """Context should NOT be numbers (unless year)"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test"
        
        # Layout: Number in B5, Text in A5, Target in C5
        ws['A5'] = 'Assets'  # Text label (ACCEPT)
        ws['B5'] = 12345  # Number (REJECT)
        ws['C5'] = 1000  # Target
        
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)
        
        model = self.parser.parse(file_obj, 'test.xlsx')
        
        # Get context for C5
        row_label, col_label = self.analyzer._get_context_labels('Test', 'C5', model.cells)
        
        print(f"\n✓ Context for C5:")
        print(f"  Row Label: {row_label}")
        
        # Should find A5 (Assets), NOT B5 (12345)
        assert row_label == 'Assets', f"Should find text label, got: {row_label}"
        assert row_label != '12345', "Should NOT return number as context"
        
        print("✓ PASS: Number rejected, text label found")
    
    def test_accept_year_as_context(self):
        """Years (2020-2030) should be accepted as context"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test"
        
        # Layout: Year in B5, Target in C5
        ws['B5'] = 2025  # Year (ACCEPT)
        ws['C5'] = 1000  # Target
        
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)
        
        model = self.parser.parse(file_obj, 'test.xlsx')
        
        # Get context for C5
        row_label, col_label = self.analyzer._get_context_labels('Test', 'C5', model.cells)
        
        print(f"\n✓ Context for C5:")
        print(f"  Row Label: {row_label}")
        
        # Should accept 2025 as a year
        assert row_label == '2025', f"Should accept year, got: {row_label}"
        
        print("✓ PASS: Year accepted as context")
    
    def test_two_column_layout(self):
        """Test 2-column layout (Assets | Liabilities)"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BS"
        
        # 2-column layout
        ws['A1'] = 'Assets'
        ws['C1'] = 'Liabilities'
        
        ws['A2'] = 'Cash'
        ws['B2'] = 1000
        
        ws['C2'] = 'Debt'
        ws['D2'] = 500
        
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)
        
        model = self.parser.parse(file_obj, 'test.xlsx')
        
        # Get context for B2 (left side)
        row_label_b2, _ = self.analyzer._get_context_labels('BS', 'B2', model.cells)
        
        # Get context for D2 (right side)
        row_label_d2, _ = self.analyzer._get_context_labels('BS', 'D2', model.cells)
        
        print(f"\n✓ 2-Column Layout:")
        print(f"  B2 Context: {row_label_b2}")
        print(f"  D2 Context: {row_label_d2}")
        
        # Should find correct labels
        assert row_label_b2 == 'Cash', f"Should find 'Cash', got: {row_label_b2}"
        assert row_label_d2 == 'Debt', f"Should find 'Debt', got: {row_label_d2}"
        
        print("✓ PASS: 2-column layout supported")
    
    def test_vietnam_plan_scenario(self):
        """Test the exact Vietnam Plan scenario"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vietnam"
        
        # Simulate Vietnam Plan layout
        ws['A4'] = 'Exchange Rate'  # Text label
        ws['B4'] = '=D2*E2'  # Formula (should be skipped)
        ws['C4'] = 'JPY/VND'  # Text label
        ws['D4'] = 201.26  # Number (should be skipped)
        ws['E4'] = 'Rate'  # Text label
        ws['F4'] = 201.26  # Target cell with hardcode
        
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)
        
        model = self.parser.parse(file_obj, 'vietnam.xlsx')
        
        # Get context for F4
        row_label, _ = self.analyzer._get_context_labels('Vietnam', 'F4', model.cells)
        
        print(f"\n✓ Vietnam Plan Scenario:")
        print(f"  F4 Context: {row_label}")
        print(f"  Cells scanned: A4='Exchange Rate', B4='=D2*E2', C4='JPY/VND', D4=201.26, E4='Rate'")
        
        # Should find E4 ('Rate'), NOT B4 (formula) or D4 (number)
        assert row_label in ['Rate', 'JPY/VND', 'Exchange Rate'], f"Should find text label, got: {row_label}"
        assert row_label != '=D2*E2', "Should NOT return formula"
        assert row_label != '201.26', "Should NOT return number"
        
        print(f"✓ PASS: Found text label '{row_label}', rejected formula and number")


if __name__ == '__main__':
    pytest.main([__file__, '-v', '-s'])
