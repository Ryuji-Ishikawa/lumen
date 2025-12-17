"""
Composite Key Matching Test Suite

Tests the Monthly Guardian's ability to match rows intelligently
even when rows are inserted, deleted, or reordered.

This is our competitive moat for monthly variance analysis.
"""

import pytest
import openpyxl
from pathlib import Path
from io import BytesIO
import sys

sys.path.insert(0, str(Path(__file__).parent.parent))
from src.parser import ExcelParser
from src.analyzer import ModelAnalyzer
from src.diff import DiffEngine


class TestCompositeKeyMatching:
    """Test suite for composite key matching"""
    
    def setup_method(self):
        """Setup for each test"""
        self.parser = ExcelParser()
        self.analyzer = ModelAnalyzer()
        self.diff_engine = DiffEngine()
    
    def create_test_file_old(self):
        """Create old version of test file"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PL"
        
        # Headers
        ws['A1'] = '勘定科目'  # Account Name
        ws['B1'] = '予算'      # Budget
        ws['C1'] = '実績'      # Actual
        
        # Data rows
        ws['A2'] = '売上高'    # Revenue
        ws['B2'] = 10000000
        ws['C2'] = '=B2*1.1'
        
        ws['A3'] = '売上原価'  # COGS
        ws['B3'] = 5000000
        ws['C3'] = '=B3*1.05'
        
        ws['A4'] = '販売費'    # Sales Expense
        ws['B4'] = 2000000
        ws['C4'] = '=B4*0.95'
        
        ws['A5'] = '営業利益'  # Operating Profit
        ws['B5'] = '=B2-B3-B4'
        ws['C5'] = '=C2-C3-C4'
        
        # Save to BytesIO
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)
        return file_obj
    
    def create_test_file_new_with_insertion(self):
        """Create new version with row inserted"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PL"
        
        # Headers
        ws['A1'] = '勘定科目'
        ws['B1'] = '予算'
        ws['C1'] = '実績'
        
        # Data rows - NEW ROW INSERTED at row 2
        ws['A2'] = '売上高'
        ws['B2'] = 10000000
        ws['C2'] = '=B2*1.1'
        
        ws['A3'] = '新規項目'  # NEW ITEM INSERTED
        ws['B3'] = 1000000
        ws['C3'] = '=B3*1.2'
        
        ws['A4'] = '売上原価'  # This moved from row 3 to row 4
        ws['B4'] = 5000000
        ws['C4'] = '=B4*1.05'
        
        ws['A5'] = '販売費'    # This moved from row 4 to row 5
        ws['B5'] = 2000000
        ws['C5'] = '=B5*0.95'
        
        ws['A6'] = '営業利益'  # This moved from row 5 to row 6
        ws['B6'] = '=B2-B4-B5'  # Formula updated to reflect new row numbers
        ws['C6'] = '=C2-C4-C5'
        
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)
        return file_obj
    
    def test_composite_key_generation(self):
        """Test that composite keys are generated correctly"""
        old_file = self.create_test_file_old()
        old_model = self.parser.parse(old_file, 'old.xlsx')
        
        # Build composite keys using column A
        keys = self.diff_engine.build_composite_keys(old_model, ['A'], 'PL')
        
        # Should have 4 keys (rows 2-5, excluding header)
        assert len(keys) >= 4, f"Should have at least 4 keys, got {len(keys)}"
        
        # Check specific keys
        assert '売上高' in keys, "Should have key for 売上高"
        assert '売上原価' in keys, "Should have key for 売上原価"
        assert '販売費' in keys, "Should have key for 販売費"
        assert '営業利益' in keys, "Should have key for 営業利益"
        
        print(f"✓ Generated {len(keys)} composite keys")
        for key, composite in list(keys.items())[:5]:
            print(f"  - Row {composite.row_number}: {key}")
    
    def test_row_matching_with_insertion(self):
        """Test that rows are matched correctly even when a row is inserted"""
        old_file = self.create_test_file_old()
        new_file = self.create_test_file_new_with_insertion()
        
        old_model = self.parser.parse(old_file, 'old.xlsx')
        new_model = self.parser.parse(new_file, 'new.xlsx')
        
        # Match rows using column A as key
        row_mapping = self.diff_engine._match_rows_by_composite_key(
            old_model, new_model, ['A'], 'PL'
        )
        
        print(f"\n✓ Row Mapping (old_row -> new_row):")
        for old_row, new_row in sorted(row_mapping.items()):
            print(f"  Row {old_row} -> Row {new_row}")
        
        # Verify correct matching
        # Old row 2 (売上高) should match new row 2
        assert row_mapping.get(2) == 2, "売上高 should match (row 2 -> row 2)"
        
        # Old row 3 (売上原価) should match new row 4 (moved due to insertion)
        assert row_mapping.get(3) == 4, "売上原価 should match (row 3 -> row 4)"
        
        # Old row 4 (販売費) should match new row 5
        assert row_mapping.get(4) == 5, "販売費 should match (row 4 -> row 5)"
        
        # Old row 5 (営業利益) should match new row 6
        assert row_mapping.get(5) == 6, "営業利益 should match (row 5 -> row 6)"
        
        print("✓ All rows matched correctly despite insertion!")
    
    def test_uniqueness_validation(self):
        """Test that uniqueness validation detects duplicate keys"""
        old_file = self.create_test_file_old()
        old_model = self.parser.parse(old_file, 'old.xlsx')
        
        # Validate keys with column A (should be unique)
        uniqueness_rate, duplicates = self.diff_engine.validate_key_uniqueness(old_model, ['A'], 'PL')
        
        print(f"\n✓ Uniqueness Rate: {uniqueness_rate*100:.1f}%")
        print(f"✓ Duplicates: {len(duplicates)}")
        
        # Should be 100% unique
        assert uniqueness_rate == 1.0, "Keys should be 100% unique"
        assert len(duplicates) == 0, "Should have no duplicates"
    
    def test_logic_change_detection(self):
        """Test that logic changes are detected correctly"""
        old_file = self.create_test_file_old()
        new_file = self.create_test_file_new_with_insertion()
        
        old_model = self.parser.parse(old_file, 'old.xlsx')
        new_model = self.parser.parse(new_file, 'new.xlsx')
        
        # Analyze both
        old_model = self.analyzer.analyze(old_model)
        new_model = self.analyzer.analyze(new_model)
        
        # Run diff with composite key matching
        diff_result = self.diff_engine.compare(old_model, new_model, ['A'], 'PL')
        
        print(f"\n✓ Logic Changes: {len(diff_result.logic_changes)}")
        print(f"✓ Input Updates: {len(diff_result.input_updates)}")
        
        # Should detect logic change in 営業利益 formula (B5 -> B6 due to row shift)
        assert len(diff_result.logic_changes) > 0, "Should detect logic changes"
        
        for change in diff_result.logic_changes[:3]:
            print(f"  - {change.description}")
    
    def test_uniqueness_validator_with_duplicates(self):
        """Test that uniqueness validator detects duplicate keys"""
        # Create file with duplicate account names
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PL"
        
        # Headers
        ws['A1'] = '勘定科目'
        ws['B1'] = '部門'  # Department
        ws['C1'] = '金額'  # Amount
        
        # Data with DUPLICATE account names
        ws['A2'] = '売上高'
        ws['B2'] = '営業部'
        ws['C2'] = 10000000
        
        ws['A3'] = '売上高'  # DUPLICATE!
        ws['B3'] = '製造部'
        ws['C3'] = 5000000
        
        ws['A4'] = '売上高'  # DUPLICATE!
        ws['B4'] = '管理部'
        ws['C4'] = 3000000
        
        ws['A5'] = '売上原価'
        ws['B5'] = '製造部'
        ws['C5'] = 2000000
        
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)
        
        model = self.parser.parse(file_obj, 'duplicate.xlsx')
        
        # Test with column A only (should have duplicates)
        uniqueness_a, duplicates_a = self.diff_engine.validate_key_uniqueness(model, ['A'], 'PL')
        
        print(f"\n✓ Column A Uniqueness: {uniqueness_a*100:.1f}%")
        print(f"✓ Duplicates found: {duplicates_a}")
        
        # Should be less than 95% unique
        assert uniqueness_a < 0.95, "Should detect duplicates in column A"
        assert len(duplicates_a) > 0, "Should have duplicate keys"
        assert '売上高' in duplicates_a, "Should identify 売上高 as duplicate"
        
        # Test with columns A+B (should be unique)
        uniqueness_ab, duplicates_ab = self.diff_engine.validate_key_uniqueness(model, ['A', 'B'], 'PL')
        
        print(f"✓ Column A+B Uniqueness: {uniqueness_ab*100:.1f}%")
        print(f"✓ Duplicates found: {duplicates_ab}")
        
        # Should be 100% unique
        assert uniqueness_ab == 1.0, "Should be unique with A+B columns"
        assert len(duplicates_ab) == 0, "Should have no duplicates with A+B"
        
        print("✓ Uniqueness validator correctly detects duplicates!")


if __name__ == '__main__':
    pytest.main([__file__, '-v', '-s'])
