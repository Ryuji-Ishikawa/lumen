"""
CSV Export Validation Test

CRITICAL: Validates that CSV export does NOT contain formulas in Context column.

UAT Failure Evidence:
- Context column contained "=(D18*E18)" and "=-D24+D25"
- This is WRONG - Context should be TEXT labels only

This test validates the entire flow:
1. Parse Excel with formulas
2. Analyze and detect risks
3. Verify Context column contains TEXT, not formulas
"""

import pytest
import sys
from pathlib import Path
from io import BytesIO
import openpyxl
import pandas as pd

sys.path.insert(0, str(Path(__file__).parent.parent))
from src.parser import ExcelParser
from src.analyzer import ModelAnalyzer


class TestCSVExportValidation:
    """Test CSV export does not contain formulas in Context"""
    
    def setup_method(self):
        """Setup for each test"""
        self.parser = ExcelParser()
        self.analyzer = ModelAnalyzer()
    
    def test_csv_context_no_formulas(self):
        """
        CRITICAL: CSV Context column should NEVER contain formulas.
        
        This test simulates the exact UAT failure scenario.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vietnam Plan"
        
        # Simulate Vietnam Plan layout with hardcoded values
        # Row 18: Text label, Formula, Hardcoded value
        ws['A18'] = '売上高'  # Text label (Sales)
        ws['B18'] = '=D18*E18'  # Formula (should be IGNORED by context)
        ws['C18'] = '=100.5'  # Hardcoded value in formula (will trigger risk)
        
        # Row 24: Text label, Formula, Hardcoded value
        ws['A24'] = '純資産'  # Text label (Net Assets)
        ws['B24'] = '=-D24+D25'  # Formula (should be IGNORED by context)
        ws['C24'] = '=200.5'  # Hardcoded value in formula (will trigger risk)
        
        # Add some driver cells so formulas are valid
        ws['D18'] = 10
        ws['E18'] = 20
        ws['D24'] = 50
        ws['D25'] = 30
        
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)
        
        # Parse and analyze
        model = self.parser.parse(file_obj, 'vietnam.xlsx')
        model = self.analyzer.analyze(model, allowed_constants=[])
        
        # Convert risks to DataFrame (simulating CSV export)
        risk_data = []
        for risk in model.risks:
            risk_data.append({
                "Risk Type": risk.risk_type,
                "Severity": risk.severity,
                "Location": risk.get_location(),
                "Context": risk.get_context(),
                "Description": risk.description
            })
        
        df = pd.DataFrame(risk_data)
        
        print("\n✓ CSV Export Preview:")
        print(df.to_string())
        
        # CRITICAL VALIDATION: Context column should NOT contain formulas
        for idx, row in df.iterrows():
            context = row['Context']
            
            print(f"\n  Row {idx}: Context = '{context}'")
            
            # REJECT: Formulas starting with =
            assert not context.startswith('='), f"Context contains formula: {context}"
            
            # REJECT: Formula patterns like =(D18*E18)
            assert '=(' not in context, f"Context contains formula pattern: {context}"
            
            # REJECT: Formula operators in suspicious patterns
            if context and len(context) > 3:
                # Allow operators in normal text, but reject formula-like patterns
                if context.startswith('=') or '=(' in context or ')' in context:
                    assert False, f"Context looks like formula: {context}"
        
        print("\n✓ PASS: No formulas found in Context column")
        
        # POSITIVE VALIDATION: Context should contain TEXT labels
        contexts = df['Context'].tolist()
        
        # Should find our text labels
        assert any('売上高' in str(c) for c in contexts), "Should find '売上高' in context"
        assert any('純資産' in str(c) for c in contexts), "Should find '純資産' in context"
        
        print("✓ PASS: Text labels found in Context column")
    
    def test_csv_export_with_numbers_rejected(self):
        """
        Test that numbers are NOT used as context labels.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test"
        
        # Layout: Number, Text, Hardcoded value
        ws['A5'] = 12345  # Number (should be IGNORED)
        ws['B5'] = 'Revenue'  # Text label (should be USED)
        ws['C5'] = '=999.5'  # Hardcoded value in formula (will trigger risk)
        
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)
        
        # Parse and analyze
        model = self.parser.parse(file_obj, 'test.xlsx')
        model = self.analyzer.analyze(model, allowed_constants=[])
        
        # Convert risks to DataFrame
        risk_data = []
        for risk in model.risks:
            risk_data.append({
                "Context": risk.get_context()
            })
        
        df = pd.DataFrame(risk_data)
        
        print("\n✓ CSV Export Preview:")
        print(df.to_string())
        
        # Validate: Should find 'Revenue', NOT '12345'
        contexts = df['Context'].tolist()
        
        assert any('Revenue' in str(c) for c in contexts), "Should find 'Revenue' in context"
        assert not any('12345' in str(c) for c in contexts), "Should NOT find '12345' in context"
        
        print("✓ PASS: Number rejected, text label used")
    
    def test_csv_export_empty_context_acceptable(self):
        """
        Test that empty context is acceptable if no text labels found.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test"
        
        # Layout: Only numbers and formulas, no text
        ws['A5'] = 100  # Number
        ws['B5'] = '=A5*2'  # Formula
        ws['C5'] = '=999.5'  # Hardcoded value in formula (will trigger risk)
        
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)
        
        # Parse and analyze
        model = self.parser.parse(file_obj, 'test.xlsx')
        model = self.analyzer.analyze(model, allowed_constants=[])
        
        # Convert risks to DataFrame
        risk_data = []
        for risk in model.risks:
            risk_data.append({
                "Context": risk.get_context()
            })
        
        df = pd.DataFrame(risk_data)
        
        print("\n✓ CSV Export Preview:")
        print(df.to_string())
        
        # Validate: Context can be empty if no text found
        contexts = df['Context'].tolist()
        
        # Empty context is acceptable
        # But should NOT contain formulas or numbers
        for context in contexts:
            if context:  # If not empty
                assert not context.startswith('='), f"Context contains formula: {context}"
                assert not context.isdigit(), f"Context is a number: {context}"
        
        print("✓ PASS: Empty context acceptable, no formulas or numbers")


if __name__ == '__main__':
    pytest.main([__file__, '-v', '-s'])
